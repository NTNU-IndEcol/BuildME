'''
Functions to convert BuildME_config.xlsx into the corresponding dictionaries
imported in settings.py
'''

import openpyxl
import pandas as pd


def read_cover(ConfigFile):
    '''
    Reads the "cover" sheet in the config file
    and returns the dictionary with simulation settings.
    Ex:
        SimulationConfig = {'cover version': 'V1.0',
                            'simulate': 'combinations',
                            'shielding': 'medium'
                            'cpus': 'auto'
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['cover']
    Cix = 2 # Column index
    Rix = 1 # Row index
    while config_sheet.cell(Rix,Cix).value != 'Simulation settings':
        Rix+=1
    Rix+=1 # First settings row index
    SimulationConfig = {}
    while config_sheet.cell(Rix,Cix).value != None:
        SimulationConfig[config_sheet.cell(Rix,Cix).value] = config_sheet.cell(Rix,Cix+1).value
        Rix+=1
    return SimulationConfig


def read_combinations(ConfigFile):
    '''
    Reads the "combinbations" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        combinations = {
            'USA':
                {'occupation': ['RT', 'MFH', 'SFH', 'HotelLarge', 'OfficeMedium', 'SchoolPrimary', 'SchoolSecondary','RetailStripmall', 'RetailStandalone'],
                 'energy standard': ['non-standard', 'standard', 'efficient', 'ZEB'],
                 'RES': ['RES0', 'RES2.1', 'RES2.2', 'RES2.1+RES2.2'],
                 'climate_region': ['1A', '2A', '2B', '3A', '3B-Coast', '3B', '3C', '4A', '4B', '4C', 5A', '5B', '6A', '6B', '7', '8'],
                 'climate_scenario': ['2015'],
                 'cooling': ['HVAC', 'MMV']
                 },
            'DE':
                {'occupation': ['RT', 'MFH', 'SFH', 'HotelLarge', 'OfficeMedium', 'SchoolPrimary', 'SchoolSecondary','RetailStripmall', 'RetailStandalone'],
                 'energy standard': ['non-standard', 'standard', 'efficient', 'ZEB'],
                 'RES': ['RES0', 'RES2.1', 'RES2.2', 'RES2.1+RES2.2'],
                 'climate_region': ['Germany'],
                 'climate_scenario': ['2015'],
                 'cooling': ['HVAC', 'MMV']
                 },
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['combinations']
    HRix = 3 # Headers row index
    HCix = 3 # Headers column index
    Rix = 4 # start row index
    Cix = 3 # start column index
    if config_sheet.cell(Rix-1, Cix).value != 'region':
        raise Exception('"region" has to be the first aspect in the "combinations" sheet of BuildME config file.')
    combinations  = {}
    while config_sheet.cell(Rix,Cix).value != None:
        region = config_sheet.cell(Rix,Cix).value
        combinations[region] = {}
        Cix+=1
        while config_sheet.cell(HRix,Cix).value != None:
            aspect = config_sheet.cell(HRix,Cix).value
            combinations[region][aspect] = eval(config_sheet.cell(Rix,Cix).value)
            Cix+=1
        Rix+=1
        Cix=HCix
    return combinations


def read_debug_combinations(ConfigFile):
    '''
    Reads the "debug combinbations" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        debug_combinations = {
            'IT':
                {'occupation': ['SFH'],
                 'energy standard': ['standard'],
                 'RES': ['RES0'],
                 'climate_region': ['Italy'],
                 'climate_scenario': ['2015'],
                 'cooling': ['HVAC']
                 }
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['debug combinations']
    HRix = 3 # Headers row index
    HCix = 3 # Headers column index
    Rix = 4 # start row index
    Cix = 3 # start column index
    debug_combinations  = {}
    if config_sheet.cell(Rix-1,Cix).value != 'region':
        raise Exception('"region" has to be the first aspect in the "debug combinations" sheet of BuildME config file.')
    while config_sheet.cell(Rix,Cix).value != None:
        region = config_sheet.cell(Rix,Cix).value
        debug_combinations[region] = {}
        Cix+=1
        while config_sheet.cell(HRix,Cix).value != None:
            aspect = config_sheet.cell(HRix,Cix).value
            debug_combinations[region][aspect] = eval(config_sheet.cell(Rix,Cix).value)
            Cix+=1
        Rix+=1
        Cix=HCix
    return debug_combinations


def read_archetype_proxies(ConfigFile):
    '''
    Reads the "archetype proxies" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        archetype_proxies = {
            ('DE', 'HotelLarge'): ('USA', 'HotelLarge'),
            ('DE', 'SFH'): ('USA', 'SFH'),
            ('CN', 'SFH'): ('USA', 'SFH') 
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['archetype proxies']
    Rix = 4 # start row index
    Cix = 3 # start column index
    archetype_proxies  = {}
    while config_sheet.cell(Rix,Cix).value != None:
        archetype = eval(config_sheet.cell(Rix,Cix).value)
        archetype_proxies[archetype] = eval(config_sheet.cell(Rix,Cix+1).value)
        Rix+=1
    return archetype_proxies


def read_climate_stations(ConfigFile):
    '''
    Reads the "climate stations" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        climate_stations = {
            'DE': {'Germany': 'Frankfurt_am_Main_Airp_-hour.epw'},
            'CN': {'I': 'CN-Harbin.epw',
                   'II': 'CN-Beijing.epw',
                   'III': 'CN-Wuhan.epw',
                   'IV': 'CN-Haikou.epw',
                   'V': 'CN-Kunming.epw'}
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['climate stations']
    Rix = 4 # start row index
    Cix = 3 # start column index
    climate_stations = {}
    while config_sheet.cell(Rix,Cix).value != None: 
        region = config_sheet.cell(Rix,Cix).value
        climate_zone = str(config_sheet.cell(Rix,Cix+1).value) # str required for those climate zones which are integers (Ex: USA, 7)
        try:
            climate_stations[region][climate_zone] = config_sheet.cell(Rix,Cix+2).value
        except:
            climate_stations[region] = {}
            climate_stations[region][climate_zone] = config_sheet.cell(Rix,Cix+2).value
        Rix+=1
    return climate_stations


def read_material_aggregation(ConfigFile):
    '''
    Reads the "material aggregation" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        material_aggregation = {'Asphalt_shingle': 'other',
                          'Air_4_in_vert': 'other',
                          'Bldg_paper_felt': 'paper and cardboard',
                          'Std Wood 6inch': 'wood and wood products'
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['material aggregation']
    Rix = 4 # start row index
    Cix = 3 # start column index
    material_aggregation = {}
    while config_sheet.cell(Rix,Cix).value != None:
        material_aggregation[ config_sheet.cell(Rix,Cix).value ] = config_sheet.cell(Rix,Cix+1).value
        Rix+=1
    return material_aggregation


def read_atypical_materials(ConfigFile):
    '''
    Reads the "atypical materials" sheet in the config file
    and returns the corresponding dictionary.
    Ex:
        climate_stations = {
            'DE': {'Germany': 'Frankfurt_am_Main_Airp_-hour.epw'},
            'CN': {'I': 'CN-Harbin.epw',
                   'II': 'CN-Beijing.epw',
                   'III': 'CN-Wuhan.epw',
                   'IV': 'CN-Haikou.epw',
                   'V': 'CN-Kunming.epw'}
            }
    '''
    config_sheet = openpyxl.load_workbook(ConfigFile)['atypical materials']
    Rix = 4 # start row index
    Cix = 3 # start column index
    atypical_materials = {}
    while config_sheet.cell(Rix,Cix).value != None: 
        material_name = config_sheet.cell(Rix,Cix).value
        atypical_materials[material_name] = {}
        atypical_materials[material_name]['density'] = config_sheet.cell(Rix,Cix+1).value
        atypical_materials[material_name]['thickness'] = config_sheet.cell(Rix,Cix+2).value
        Rix+=1
    return atypical_materials


def read_climate_region_weight(ConfigFile):
    '''
    Reads the "climate region weight" sheet in the config file
    and returns the corresponding dataframe.
    '''
    df = pd.read_excel(ConfigFile, 'climate region weight', header=2)
    df = df.dropna(how='all').dropna(how='all', axis=1)
    return df


def read_surrogate_elements(ConfigFile, combinations):
    '''
    Reads the "surrogate elements" sheet in the config file
    and returns the corresponding dataframe.
    '''
    df = pd.read_excel(ConfigFile, 'surrogate elements', header=2)
    df = df.dropna(how='all').dropna(how='all', axis=1)
    df = df.fillna(value=0)
    df = df.drop(columns=['comment'])
    """aspect_names = ['region'] + list(list(combinations.values())[0].keys())
    where_surrogate = list(df.columns).index('surrogate')
    cols_to_idx = [v for i, v in enumerate(list(df.columns)) if i < where_surrogate]
    unknown_items = []
    for item in cols_to_idx:
        if item not in aspect_names:
            unknown_items.append(item)
        if unknown_items:
            raise Exception(f'Variable surrogate_elements contains unknown column name(s): {unknown_items}')
    df = df.set_index(cols_to_idx)
    df = df.sort_index()"""
    return df


def read_config_file(ConfigFile):
    '''
    Reads the config files and returns the dictionaries defined in the settings,
    using the functions defined above.

    '''
    SimulationConfig       = read_cover(ConfigFile)    
    combinations           = read_combinations(ConfigFile)
    debug_combinations     = read_debug_combinations(ConfigFile)
    archetype_proxies      = read_archetype_proxies(ConfigFile)   
    climate_stations       = read_climate_stations(ConfigFile)  
    material_aggregation   = read_material_aggregation(ConfigFile)
    atypical_materials     = read_atypical_materials(ConfigFile)
    climate_region_weight = read_climate_region_weight(ConfigFile)
    surrogate_elements = read_surrogate_elements(ConfigFile, combinations)
    return SimulationConfig, combinations, debug_combinations, archetype_proxies, climate_stations, \
           material_aggregation, atypical_materials, climate_region_weight, surrogate_elements


