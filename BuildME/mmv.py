"""
Functions to perform implementation of mixed-mode ventilation (MMV) i.e. cooling dependent on thermal comfort

Copyright: Kamila Krych, 2021
"""
from eppy.modeleditor import IDF
import pandas as pd
import numpy as np
import openpyxl
from . import settings
import os
import datetime


def change_archetype_to_MMV(idf, occupation, xlsx_mmv, dir_replace_mmv):
    """
    Converts an idf file to one with mixed mode ventilation (MMV) i.e. cooling both through HVAC and window opening
    :param idf: The .idf file
    :param occupation: the occupation of the chosen archetype (e.g. SFH)
    :param xlsx_mmv: Path and filename of the mmv-implementation.xlsx file, normally './data/mmv-implementation.xlsx'
    :param dir_replace_mmv: Path and filename of the replace-mmv.xlsx file, normally './data/replace_mmv.xlsx'
    :return idf: The .idf file with MMV implemented
    """
    shielding = settings.shielding
    # Create dictionaries needed for the MMV procedure
    surface_dict = create_surface_dict(idf)  # create surfaces for which AFN is suitable
    zone_dict_mmv, zone_dict_non_mmv, surface_dict = create_zone_dicts(idf, surface_dict, occupation)
    window_dict = create_window_dict(idf)
    surfaces_f_dict, surfaces_nf_dict = create_surface_group_dicts(surface_dict)
    # Modify the idf file
    idf = create_WPC_curves(idf, zone_dict_mmv, shielding)
    idf = change_idf_objects(idf, xlsx_mmv, zone_dict_mmv)
    idf = create_idf_objects(idf, xlsx_mmv, zone_dict_mmv, zone_dict_non_mmv, window_dict,
                             surface_dict, surfaces_f_dict, surfaces_nf_dict, shielding)
    idf = delete_idf_objects(idf, xlsx_mmv)
    idf = write_EMS_program(idf, xlsx_mmv, zone_dict_mmv, window_dict)
    # Update the replace.xlsx file (only if necessary)
    create_or_update_excel_replace(occupation, xlsx_mmv, surfaces_f_dict, surfaces_nf_dict, dir_replace_mmv)
    # idf.saveas('./files/with-MMV.idf')
    return idf


def load_xlsx_data(in_file, sheet_name):
    """
    Loads the data from an .xlsx file and returns them in a pandas dataframe
    :param in_file: The .xlsx file directory
    :param sheet_name: The name of the sheet in the .xlsx file
    :return df: Pandas dataframe with the contents of the sheet
    """
    df = pd.read_excel(in_file, sheet_name=sheet_name, engine='openpyxl')
    return df


def create_zone_dicts(idf, surface_dict, occupation):
    """
    Creates two dictionaries with zone information, both with integer keys
    :param idf: The .idf file
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :param occupation: Building archetype
    :return zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :return zone_dict_non_mmv: Dictionary of other zones (without HVAC or without people)
    """
    object_types = ['HVACTemplate:Zone:IdealLoadsAirSystem', 'ZoneHVAC:IdealLoadsAirSystem']
    zone_dict_mmv = {}
    zone_dict_non_mmv = {}
    i = 1
    j = 1
    hvac_zone_objs = [obj for obj_type in object_types for obj in idf.idfobjects[obj_type]]
    assert hvac_zone_objs, "There are no 'IdealLoadsAirSystem' objects in '%s'. Mixed mode ventilation MMV " \
                           "can only be added if the HVAC system is implemented through 'IdealLoads'." % occupation
    for zone in idf.idfobjects['Zone']:
        surfaces_for_zone = [k for k, v in surface_dict.items() if v['Zone'] == zone.Name]
        if len(surfaces_for_zone) > 0:
            hvac_zone = [obj for obj in hvac_zone_objs if obj.Zone_Name == zone.Name]
            if hvac_zone:
                obj = hvac_zone[0]
                zone_dict_mmv[i] = {}
                zone_dict_mmv[i]['Zone_Name'] = obj.Zone_Name
                zone_dict_mmv[i]['Template_Thermostat_Name'] = obj.Template_Thermostat_Name
                people_obj = [obj2.Name for obj2 in idf.idfobjects['People'] if
                              obj2.Zone_or_ZoneList_Name == obj.Zone_Name]
                if people_obj:
                    zone_dict_mmv[i]['People'] = people_obj[0]
                    i += 1
                    for k in surfaces_for_zone:  # add a flag that this surface is in a MMV zone
                        surface_dict[k]['is_in_MMV_zone'] = True
                else:  # if no people in the zone, MMV cannot be implemented -> zone is treated as AFN (non-hvac) zone
                    zone_dict_mmv.pop(i)
                    zone_dict_non_mmv[j] = {}
                    zone_dict_non_mmv[j]['Zone_Name'] = zone.Name
                    j += 1
            else:  # if no hvac objects associated with the zone
                zone_dict_non_mmv[j] = {}
                zone_dict_non_mmv[j]['Zone_Name'] = zone.Name
                j += 1
        else:  # if no surfaces associated with the zone
            zone_dict_non_mmv[j] = {}
            zone_dict_non_mmv[j]['Zone_Name'] = zone.Name
            j += 1
    return zone_dict_mmv, zone_dict_non_mmv, surface_dict


def create_window_dict(idf):
    """
    Creates a dictionary with window information, with integer keys
    :param idf: The .idf file
    :return window_dict: Dictionary of windows (their names, and surface and zone they belong to)
    """
    idf_object = 'Window'
    window_dict = {}
    i = 1
    for obj in idf.idfobjects[idf_object]:
        window_dict[i] = {}
        window_dict[i]['Name'] = obj.Name
        surface = obj.Building_Surface_Name
        window_dict[i]['Surface'] = surface
        zone = [obj2.Zone_Name for obj2 in idf.idfobjects['BuildingSurface:Detailed'] if obj2.Name == surface]
        window_dict[i]['Zone'] = zone[0]
        i += 1
    idf_object = 'FenestrationSurface:Detailed'
    for obj in idf.idfobjects[idf_object]:
        if obj.Surface_Type == 'Window':
            window_dict[i] = {}
            window_dict[i]['Name'] = obj.Name
            surface = obj.Building_Surface_Name
            window_dict[i]['Surface'] = surface
            zone = [obj2.Zone_Name for obj2 in idf.idfobjects['BuildingSurface:Detailed'] if obj2.Name == surface]
            window_dict[i]['Zone'] = zone[0]
            i += 1
    return window_dict


def create_surface_dict(idf):
    """
    Creates a dictionary with surface information (walls, floors, windows etc.), with integer keys. Only includes
    surfaces for which AFN model can be implemented (doesn't include GroundFCfactorMethod objects, internal partitions,
    and surfaces like ceilings for which we don't have AFN infiltration values)
    :param idf: The .idf file
    :return surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    """
    surface_dict = {}
    # zone_list = [k['Zone_Name'] for v, k in zone_dict.items()]
    i = 1
    for idf_object in ['BuildingSurface:Detailed']:
        for obj in idf.idfobjects[idf_object]:
            # if obj.Zone_Name in zone_list:  # only include surfaces that belong to an HVAC zone
            outside = obj.Outside_Boundary_Condition
            surface_type = obj.Surface_Type
            if obj.Name == obj.Outside_Boundary_Condition_Object:
                # the same name indicates internal partition separating like zones
                continue
            if outside == 'Outdoors' and surface_type == 'Wall':
                surface_group = 'Walls external'
            elif outside == 'GroundFCfactorMethod' and surface_type == 'Wall':
                # quoting an E+ error: "This type of surface (has ground, etc exposure) cannot be used in the AiflowNetwork model."
                continue
            elif surface_type == 'Wall':
                surface_group = 'Walls internal'
            elif surface_type == 'Roof':
                surface_group = 'Roofs'
            elif outside == 'Surface' and surface_type == 'Floor':
                surface_group = 'Floors internal'
            elif outside == 'Outdoors' and surface_type == 'Floor':
                surface_group = 'Floors external'
            else:
                continue  # ignores surfaces such as ceilings, as we don't have AFN infiltration values for these
            surface_dict[i] = {}
            surface_dict[i]['Name'] = obj.Name
            surface_dict[i]['Object_Type'] = idf_object
            surface_dict[i]['Surface_Group'] = surface_group
            surface_dict[i]['Area'] = calculate_area(obj)
            surface_dict[i]['Zone'] = obj.Zone_Name
            surface_dict[i]['is_in_MMV_zone'] = False  # default value
            i += 1
    for idf_object in ['Window', 'Door', 'FenestrationSurface:Detailed']:
        for obj in idf.idfobjects[idf_object]:
            wall_obj = get_wall_object_from_fenestration(idf, obj)
            if wall_obj.Outside_Boundary_Condition == 'Outdoors':
                if check_if_window(idf_object, obj):
                    surface_group = 'Windows external'
                else:
                    surface_group = 'Doors external'
            else:
                if check_if_window(idf_object, obj):
                    surface_group = 'Windows internal'
                else:
                    surface_group = 'Doors internal'
            surface_dict[i] = {}
            surface_dict[i]['Name'] = obj.Name
            surface_dict[i]['Object_Type'] = idf_object
            surface_dict[i]['Surface_Group'] = surface_group
            surface_dict[i]['Area'] = calculate_area(obj)
            surface = obj.Building_Surface_Name
            zone = [obj2.Zone_Name for obj2 in idf.idfobjects['BuildingSurface:Detailed'] if obj2.Name == surface]
            surface_dict[i]['Zone'] = zone[0]
            surface_dict[i]['is_in_MMV_zone'] = False  # default value
            i += 1
    return surface_dict


def check_if_window(idf_object, obj):
    """
    Checks if a given object is a window
    :param idf_object: Idf object name
    :param obj: Idf object
    :return window: Boolean telling if a given object is a window
    """
    if idf_object == 'Window':
        window = True
    elif idf_object == 'FenestrationSurface:Detailed':
        if obj.Surface_Type == 'Window':
            window = True
        else:
            window = False
    else:
        window = False
    return window


def create_surface_group_dicts(surface_dict):
    """
    Creates a dictionary grouping surfaces from the same surface group and of similar area, with integer keys
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :return surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :return surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    """
    surfaces_f_dict = {}
    surfaces_nf_dict = {}
    a1 = np.array([surface_number['Surface_Group'] for surface_number in surface_dict.values()])
    a2 = np.array([round(surface_number['Area'], ndigits=2) for surface_number in surface_dict.values()])
    ar = np.column_stack((a1, a2))
    unique, unique_inverse = np.unique(ar, axis=0, return_inverse=True)  # find unique surface_group&area combinations
    k1 = 1
    k2 = 1
    for i in range(0, np.shape(unique)[0]):
        surface_group = unique[i, 0]
        if surface_group in ['Doors external', 'Doors internal', 'Windows external', 'Windows internal']:
            surfaces_f_dict = fill_surface_group_dict(i, k1, surfaces_f_dict, unique, unique_inverse, surface_dict)
            k1 += 1
        else:  # ['Walls external', 'Walls internal', 'Floors external', 'Roofs']
            surfaces_nf_dict = fill_surface_group_dict(i, k2, surfaces_nf_dict, unique, unique_inverse, surface_dict)
            k2 += 1
    return surfaces_f_dict, surfaces_nf_dict


def fill_surface_group_dict(i, k, dic, unique, unique_inverse, surface_dict):
    """
    Fills a surface group dictionary with given values
    :param i: Iteration (through a number of unique surface_group & area combinations)
    :param k: Integer key
    :param dic: Dictionary of surfaces (area, surface group, list of surfaces' names)
    :param unique: The sorted unique values (surface_group & area)
    :param unique_inverse: The indices to reconstruct the original array from the unique array
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :return dic: Dictionary of surfaces (area, surface group, list of surfaces' names) with additional entry
    """
    dic[k] = {}
    dic[k]['Surface_Group'] = unique[i, 0]
    dic[k]['Area'] = unique[i, 1]
    dic[k]['Names'] = []
    for j in range(0, np.shape(unique_inverse)[0]):
        if unique_inverse[j] == i:  # find all the surfaces that belong to this group
            name = surface_dict[j + 1]['Name']
            dic[k]['Names'].append(name)
    return dic


def change_idf_objects(idf, xlsx_mmv, zone_dict_mmv):
    """
    Takes an idf file and changes some of its objects according to the MMV procedure instructions
    :param idf: The .idf file
    :param xlsx_mmv: The .xlsx file with MMV procedure instructions
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :return idf: The .idf file with changed objects
    """
    df = load_xlsx_data(xlsx_mmv, 'change')
    for i in range(0, len(df.index)):
        idf_object = df.loc[i, 'idfobject']
        object_field = df.loc[i, 'objectfield']
        value = df.loc[i, 'value']
        for obj in idf.idfobjects[idf_object]:
            if value[-14:] == '<mmv zone no.>':
                zone_list = [zone_number['Zone_Name'] for zone_number in zone_dict_mmv.values()]
                if obj.Zone_Name in zone_list:
                    zone_no = list(zone_dict_mmv.keys())[zone_list.index(obj.Zone_Name)]
                    obj[object_field] = value[:-14] + str(zone_no)
            else:
                obj[object_field] = value
    return idf


def create_idf_objects(idf, xlsx_mmv, zone_dict_mmv, zone_dict_non_mmv, window_dict, surface_dict,
                       surfaces_f_dict, surfaces_nf_dict, shielding):
    """
    Takes an idf file and creates some new objects according to the MMV procedure instructions
    :param idf: The .idf file
    :param xlsx_mmv: The .xlsx file with MMV procedure instructions
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param zone_dict_non_mmv: Dictionary of other zones (without HVAC or without people)
    :param window_dict: Dictionary of windows (their names, and surface and zone they belong to)
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :param surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :param surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    :param shielding: the level of wind shielding (low, medium, high)
    :return idf:
    """
    df = load_xlsx_data(xlsx_mmv, 'create')
    for i in df['object number'].unique():  # each object to be created
        df_i = df[df['object number'] == i]
        df_i = df_i.reset_index()
        idf_object = df_i.loc[0, 'idfobject']
        j = 1
        while True:  # the number of loops is variable (the number of objects of a given type to be created)
            name_in = df_i.loc[0, 'value']
            name, more_loops_j = assign_name(str(name_in), zone_dict_mmv, zone_dict_non_mmv, window_dict,
                                             surface_dict, surfaces_f_dict, surfaces_nf_dict, j)
            if name is None:
                break
            if name != '':
                new_object = idf.newidfobject(idf_object)
                name_field = df_i.loc[0, 'objectfield']
                new_object[name_field] = name
            else:
                j += 1
                continue
            for k in range(1, len(df_i.index)):  # each field to be filled in (except for the name field)
                object_field = df_i.loc[k, 'objectfield']
                value_in = df_i.loc[k, 'value']
                new_object[object_field] = assign_value(str(value_in), zone_dict_mmv, window_dict, surface_dict,
                                                        surfaces_f_dict, surfaces_nf_dict, j, idf, shielding)
            j += 1
            if not more_loops_j:  # if there are no more objects of this type to be created
                break
    return idf


def delete_idf_objects(idf, xlsx_mmv):
    df = load_xlsx_data(xlsx_mmv, 'delete')
    for idf_object in df['idfobject']:
        objs = idf.idfobjects[idf_object]
        while objs:
            idf.removeidfobject(objs[0])
    return idf


def assign_name(name_in, zone_dict_mmv, zone_dict_non_mmv, window_dict, surface_dict,
                surfaces_f_dict, surfaces_nf_dict, it):
    """
    Assigns name for new idf object based on input name
    :param name_in: Input idf object name
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param zone_dict_non_mmv: Dictionary of other zones (without HVAC or without people)
    :param window_dict: Dictionary of windows (their names, and surface and zone they belong to)
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :param surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :param surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    :param it: Iteration
    :return name: Final idf object name
    :return more_loops: Boolean telling if more objects of a given type are to be created
    """
    # all names with <...> need to be customized
    if name_in == '<mmv zone name>':
        it_dict = zone_dict_mmv
        if it_dict:
            name = zone_dict_mmv[it]['Zone_Name']
        else:  # if the dict is empty
            name = None
    elif name_in == '<non-mmv zone name>':
        it_dict = zone_dict_non_mmv
        if it_dict:
            name = zone_dict_non_mmv[it]['Zone_Name']
        else:  # if the dict is empty
            name = None
    elif name_in[-14:] == '<mmv zone no.>':
        it_dict = zone_dict_mmv
        if it_dict:
            name = name_in[:-14] + str(it)
        else:  # if the dict is empty
            name = None
    elif name_in[-12:] == '<window no.>':
        name = name_in[:-12] + str(it)
        it_dict = window_dict
    elif name_in[-13:] == '<surface no.>':
        it_dict = surface_dict
        external = ['Walls external', 'Windows external', 'Doors external', 'Roofs', 'Floors external']
        if surface_dict[it]['Surface_Group'] in external:
            name = name_in[:-13] + str(it)
        else:
            name = ''
    elif name_in == '<Crack_non-fen. surface type>':
        name = 'Crack_' + str(it)
        it_dict = surfaces_nf_dict
    elif name_in == '<Opening_fen. surface type>':
        name = 'Opening_' + str(it)
        it_dict = surfaces_f_dict
    elif name_in == '<surface name>':
        name = surface_dict[it]['Name']
        it_dict = surface_dict
    else:
        name = name_in  # in some cases, only one object of a given type is to be created (so no further iterations)
        it_dict = {}
    if it < len(it_dict.keys()) and name != name_in:  # checks if there are more objects of a given type to be created
        more_loops = True
    else:
        more_loops = False
    return name, more_loops


def return_value_if_not_empty(value, dictionary):
    if dictionary:
        return value
    else: # if dictionary is empty
        return None


def assign_value(value_in, zone_dict_mmv, window_dict, surface_dict, surfaces_f_dict, surfaces_nf_dict, it, idf,
                 shielding):
    """
    Assigns value for idf object field based on input value
    :param value_in: Input idf object field value
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param window_dict: Dictionary of windows (their names, and surface and zone they belong to)
    :param surface_dict: Dictionary of surfaces (their names, idf object type, surface group, area and zone name)
    :param surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :param surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    :param it: Iteration
    :param idf: The .idf file
    :param shielding: the level of wind shielding (low, medium, high)
    :return value: Final idf object field value
    """
    # all values with <...> need to be customized
    if value_in == '<mmv zone name>':
        value = zone_dict_mmv[it]['Zone_Name']
    elif value_in[-14:] == '<mmv zone no.>':
        value = value_in[:-14] + str(it)
    elif value_in == '<people object name>':
        value = zone_dict_mmv[it]['People']
    elif value_in == '<window name>':
        value = window_dict[it]['Name']
    elif value_in == '<wind pressure coefficient curve name>':
        surface_name = surface_dict[it]['Name']
        obj_type = surface_dict[it]['Object_Type']
        obj = [obj for obj in idf.idfobjects[obj_type] if obj.Name == surface_name][0]
        if obj_type == 'BuildingSurface:Detailed':
            value = assign_wpc_curve(obj, idf, zone_dict_mmv, shielding)
        else:  # if not 'BuildingSurface:Detailed', it has to be window or door
            wall_obj = get_wall_object_from_fenestration(idf, obj)
            value = assign_wpc_curve(wall_obj, idf, zone_dict_mmv, shielding)
    elif value_in[-13:] == '<surface no.>':
        # internal = ['Walls internal', 'Windows internal', 'Doors internal', 'Floors internal']
        external = ['Walls external', 'Windows external', 'Doors external', 'Roofs', 'Floors external']
        if surface_dict[it]['Surface_Group'] in external:
            value = value_in[:-13] + str(it)
        else:
            value = ''
    elif value_in == 'Crack_ or Opening_<surface type>':
        name = surface_dict[it]['Name']
        if surface_dict[it]['Object_Type'] in ['Window', 'Door', 'FenestrationSurface:Detailed']:
            for k, v in surfaces_f_dict.items():
                if name in v['Names']:
                    value = "Opening_" + str(k)
        else:
            for k, v in surfaces_nf_dict.items():
                if name in v['Names']:
                    value = "Crack_" + str(k)
    elif value_in == '<venting availability>':
        surface_group = surface_dict[it]['Surface_Group']
        if surface_group.split(' ')[0] == 'Doors':
            value = 'always_off_MMV'
        elif surface_dict[it]['is_in_MMV_zone'] is False:
            value = 'always_off_MMV'
        else:
            value = 'always_on_MMV'
    elif value_in == '<heating schedule in zone>':
        thermostat_name = zone_dict_mmv[it]['Template_Thermostat_Name']
        heat_sch = [obj.Heating_Setpoint_Schedule_Name for obj in idf.idfobjects['HVACTemplate:Thermostat'] if obj.Name == thermostat_name]
        if heat_sch:
            value = heat_sch[0]
        else:
            print(f'The heating schedule for the "{thermostat_name}" thermostat was not found')
    elif value_in == '<cooling schedule in zone>':
        thermostat_name = zone_dict_mmv[it]['Template_Thermostat_Name']
        cool_sch = [obj.Cooling_Setpoint_Schedule_Name for obj in idf.idfobjects['HVACTemplate:Thermostat'] if obj.Name == thermostat_name]
        if cool_sch:
            value = cool_sch[0]
        else:
            print(f'The heating schedule for the "{thermostat_name}" thermostat was not found')
    else:
        value = value_in
    return value


def get_wall_object_from_fenestration(idf, fenestration_obj):
    """
    Returns the name of the wall object that a given fenestration object belongs to
    :param idf: The .idf file
    :param fenestration_obj: Idf fenestration object (window or door)
    :return wall_obj: Idf wall object
    """
    wall = fenestration_obj.Building_Surface_Name
    wall_obj = [obj for obj in idf.idfobjects['BuildingSurface:Detailed'] if obj.Name == wall][0]
    return wall_obj


def assign_wpc_curve(obj, idf, zone_dict_mmv, shielding):
    """
    Assigns a wind pressure coefficient (WPC) curve according to the azimuth of the input surface object
    :param obj: Surface object
    :param idf: The .idf file
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param shielding: the level of wind shielding (low, medium, high)
    :return wpc_curve_name: Name of the WPC curve (object with wind pressure coefficient values)
    """
    deg = int(obj.azimuth)
    surface_type = obj.Surface_Type
    dirs = ['N', 'E', 'S', 'W']
    degs = [0, 90, 180, 270]  # North=0, East=90, South=180, West=270
    # (see IO reference or https://bigladdersoftware.com/epx/docs/8-0/input-output-reference/page-011.html)
    nearest_90 = 90*round(deg/90)  # round in case the wall is facing a direction off the 0-90-270-360 ones
    ind = [i for i, x in enumerate(degs) if x == nearest_90][0]
    if surface_type == 'Wall':
        wpc_curve_name = create_WPC_prefix(idf, zone_dict_mmv, shielding) + '_wall_' + dirs[ind]
    else:  # if not Wall, it has to be roof
        wpc_curve_name = create_WPC_prefix(idf, zone_dict_mmv, shielding) + '_roof_' + dirs[ind]
    return wpc_curve_name


def create_WPC_prefix(idf, zone_dict_mmv, shielding):
    """
    Creates a prefix that later helps to find the right wind pressure coefficient (WPC) curves for given wind shielding
    and building height
    :param idf: The .idf file
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param shielding: the level of wind shielding (low, medium, high)
    :return WPC_prefix: String that that later helps to find the right WPC curves
    """
    no_of_floors = estimate_no_of_floors(idf, zone_dict_mmv)
    if no_of_floors > 3:
        WPC_prefix = 'HighRise'
    else:
        WPC_prefix = 'LowRise'
    if shielding == 'medium':
        WPC_prefix = WPC_prefix + '_MedShield'
    elif shielding == 'low':
        WPC_prefix = WPC_prefix + '_LowShield'
    elif shielding == 'high':
        WPC_prefix = WPC_prefix + '_HighShield'
    else:
        raise ValueError("Shielding level '%s' is not known." % shielding)
    return WPC_prefix


def retrieve_WPC_values(name):
    """
    Find values for a given WPC curve (either degrees for "WPC array" or coefficient values for "WPC values" object)
    :param name: WPC curve name
    :return values: values for the WPC curve
    """
    LowRise_degrees = [0, 45, 90, 135, 180, 225, 270, 315]
    LowRise_MedShield_wall_N = [0.4, -0.1, -0.3, -0.35, -0.2, -0.35, -0.3, 0.1]
    LowRise_MedShield_wall_S = [-0.2, -0.35, -0.3, 0.1, 0.4, 0.1, -0.3, -0.35]
    LowRise_MedShield_wall_E = [-0.3, -0.35, -0.2, -0.35, -0.3, 0.1, 0.4, 0.1]
    LowRise_MedShield_wall_W = [-0.3, 0.1, 0.4, 0.1, -0.3, -0.35, -0.2, -0.35]
    LowRise_MedShield_roof_N = [0.3, -0.5, -0.6, -0.5, -0.5, -0.5, -0.6, -0.5]
    LowRise_MedShield_roof_S = [-0.5, -0.5, -0.6, -0.5, 0.3, -0.5, -0.6, -0.5]
    LowRise_MedShield_roof_E = [-0.6, -0.5, -0.5, -0.5, -0.6, -0.5, 0.3, -0.5]
    LowRise_MedShield_roof_W = [-0.6, -0.5, 0.3, -0.5, -0.6, -0.5, -0.5, -0.5]

    LowRise_LowShield_wall_N = [0.7, 0.35, -0.5, -0.4, -0.2, -0.4, -0.5, 0.35]
    LowRise_LowShield_wall_S = [-0.2, -0.4, -0.5, 0.35, 0.7, 0.35, -0.5, -0.4]
    LowRise_LowShield_wall_E = [-0.5, -0.4, -0.2, -0.4, -0.5, 0.35, 0.7, 0.35]
    LowRise_LowShield_wall_W = [-0.5, 0.35, 0.7, 0.35, -0.5, -0.4, -0.2, -0.4]
    LowRise_LowShield_roof_N = [0.3, -0.4, -0.6, -0.4, -0.5, -0.4, -0.6, -0.4]
    LowRise_LowShield_roof_S = [-0.5, -0.4, -0.6, -0.4, 0.3, -0.4, -0.6, -0.4]
    LowRise_LowShield_roof_E = [-0.6, -0.4, -0.5, -0.4, -0.6, -0.4, 0.3, -0.4]
    LowRise_LowShield_roof_W = [-0.6, -0.4, 0.3, -0.4, -0.6, -0.4, -0.5, -0.4]

    LowRise_HighShield_wall_N = [0.2, -0.05, -0.25, -0.3, -0.25, -0.3, -0.25, 0.05]
    LowRise_HighShield_wall_S = [-0.25, -0.3, -0.25, 0.05, 0.2, 0.05, -0.25, -0.3]
    LowRise_HighShield_wall_E = [-0.25, -0.3, -0.25, -0.3, -0.25, 0.05, 0.2, 0.05]
    LowRise_HighShield_wall_W = [-0.25, 0.05, 0.2, 0.05, -0.25, -0.3, -0.25, -0.3]
    LowRise_HighShield_roof_N = [0.25, -0.3, -0.5, -0.3, -0.4, -0.3, -0.5, -0.3]
    LowRise_HighShield_roof_S = [-0.4, -0.3, -0.5, -0.3, 0.25, -0.3, -0.5, -0.3]
    LowRise_HighShield_roof_E = [-0.5, -0.3, -0.4, -0.3, -0.5, -0.3, 0.25, -0.3]
    LowRise_HighShield_roof_W = [-0.5, -0.3, 0.25, -0.3, -0.5, -0.3, -0.4, -0.3]

    HighRise_degrees = [0, 22.5, 45, 67.5, 90, 112.5, 135, 157.5, 180, 202.5, 225, 247.5, 270, 292.5, 315, 337.5]
    HighRise_HighShield_wall_N = [0.099, 0.078, 0.042, -0.018, -0.077, -0.147, -0.145, -0.11, -0.079, -0.11,
                                  -0.145, -0.147, -0.077, -0.018, 0.042, 0.078]
    HighRise_HighShield_wall_S = [-0.079, -0.11, -0.145, -0.147, -0.077, -0.018, 0.042, 0.078, 0.099, 0.078,
                                  0.042, -0.018, -0.077, -0.147, -0.145, -0.11]
    HighRise_HighShield_wall_E = [-0.077, -0.147, -0.145, -0.11, -0.079, -0.11, -0.145, -0.147, -0.077, -0.018,
                                  0.042, 0.078, 0.099, 0.078, 0.042, -0.018]
    HighRise_HighShield_wall_W = [-0.077, -0.018, 0.042, 0.078, 0.099, 0.078, 0.042, -0.018, -0.077, -0.147,
                                  -0.145, -0.11, -0.079, -0.11, -0.145, -0.147]
    HighRise_HighShield_roof = [-0.077, -0.077, -0.077, -0.077, -0.077, -0.077, -0.077, -0.077, -0.077, -0.077,
                                -0.077, -0.077, -0.077, -0.077, -0.077]

    HighRise_LowShield_wall_N = [0.295, 0.233, 0.125, -0.053, -0.23, -0.438, -0.43, -0.328, -0.235, -0.328,
                                 -0.43, -0.438, -0.23, -0.053, 0.125, 0.233]
    HighRise_LowShield_wall_S = [-0.235, -0.328, -0.43, -0.438, -0.23, -0.053, 0.125, 0.233, 0.295, 0.233,
                                 0.125, -0.053, -0.23, -0.438, -0.43, -0.328]
    HighRise_LowShield_wall_E = [-0.23, -0.438, -0.43, -0.328, -0.235, -0.328, -0.43, -0.438, -0.23, -0.053,
                                 0.125, 0.233, 0.295, 0.233, 0.125, -0.053]
    HighRise_LowShield_wall_W = [-0.23, -0.053, 0.125, 0.233, 0.295, 0.233, 0.125, -0.053, -0.23, -0.438,
                                 -0.43, -0.328, -0.235, -0.328, -0.43, -0.438]
    HighRise_LowShield_roof = [-0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23, -0.23,
                               -0.23, -0.23, -0.23, -0.23]

    HighRise_MedShield_wall_N = [0.191, -0.149, 0.081, -0.034, -0.149, -0.283, -0.278, -0.212, -0.152, -0.212,
                                 -0.278, -0.283, -0.149, -0.034, 0.081, -0.149]
    HighRise_MedShield_wall_S = [-0.152, -0.212, -0.278, -0.283, -0.149, -0.034, 0.081, -0.149, 0.191, -0.149,
                                 0.081, -0.034, -0.149, -0.283, -0.278, -0.212]
    HighRise_MedShield_wall_E = [-0.149, -0.283, -0.278, -0.212, -0.152, -0.212, -0.278, -0.283, -0.149, -0.034,
                                 0.081, -0.149, 0.191, -0.149, 0.081, -0.034]
    HighRise_MedShield_wall_W = [-0.149, -0.034, 0.081, -0.149, 0.191, -0.149, 0.081, -0.034, -0.149, -0.283,
                                 -0.278, -0.212, -0.152, -0.212, -0.278, -0.283]
    HighRise_MedShield_roof = [-0.149, -0.149, -0.149, -0.149, -0.149, -0.149, -0.149, -0.149, -0.149, -0.149, -0.149,
                               -0.149, -0.149, -0.149, -0.149, -0.149]
    try:
        values = locals()[name]
    except KeyError:
        values = locals()[name[:-2]]  # in case we have just "..._roof" and no e.g. "..._roof_N"
    return values


def create_WPC_curves(idf, zone_dict_mmv, shielding):
    """
    Create wind pressure coefficient (WPC) curves for a given wind shielding level
    :param idf: The .idf file
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param shielding: the level of wind shielding (low, medium, high)
    :return idf: The .idf file with WPC curve objects included
    """
    WPC_prefix = create_WPC_prefix(idf, zone_dict_mmv, shielding)
    new_object = idf.newidfobject('AirflowNetwork:MultiZone:WindPressureCoefficientArray')
    rise = WPC_prefix.split("_")[0]
    name_deg = rise + '_degrees'
    values = retrieve_WPC_values(name_deg)
    new_object['Name'] = name_deg
    for i in range(0, len(values)):  # fill the WPC Array with values of degrees
        new_object['Wind_Direction_' + str(i + 1)] = values[i]
    for s in ['_wall_N', '_wall_S', '_wall_E', '_wall_W', '_roof_N', '_roof_S', '_roof_E', '_roof_W']:
        new_object = idf.newidfobject('AirflowNetwork:MultiZone:WindPressureCoefficientValues')
        name = WPC_prefix + s
        values = retrieve_WPC_values(name)
        new_object['Name'] = name
        new_object['AirflowNetworkMultiZoneWindPressureCoefficientArray_Name'] = name_deg
        for i in range(0, len(values)):  # fill the WPC curve with values of coefficients
            new_object['Wind_Pressure_Coefficient_Value_' + str(i + 1)] = values[i]
    return idf


def calculate_area(obj):
    """
    Calculates the surface object area
    :param obj: Idf surface object
    :return area: Area of the idf surface object
    """
    try:
        area = obj.area
    except:
        area = obj.Length * obj.Height
    return area


def write_EMS_program(idf, xlsx_mmv, zone_dict_mmv, window_dict):
    """
    Creates idf objects of type EnergyManagementSystem:Program according to the MMV procedure
    :param idf: The .idf file
    :param xlsx_mmv: The .xlsx file with MMV procedure instructions
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :param window_dict: Dictionary of windows (their names, and surface and zone they belong to)
    :return idf: The .idf file with EMS program
    """
    df = load_xlsx_data(xlsx_mmv, 'EMS program')
    i = 1
    while True:
        new_object = idf.newidfobject('EnergyManagementSystem:Program')
        name_field = df.loc[0, 'objectfield']
        name = "DefineOpening_" + str(i)
        if i < len(zone_dict_mmv.keys()):
            more_loops_i = True
        else:
            more_loops_i = False
        new_object[name_field] = name
        k = 1
        for j in range(1, len(df.index)):  # each field to be filled in (except for the name field)
            value_in = df.loc[j, 'value']
            value_in = value_in.replace('X', str(i))
            if value_in[:21] == 'SET Control_<windows>':
                window_dict_zone = [key for key, v in window_dict.items() if
                                    v['Zone'] == zone_dict_mmv[i]['Zone_Name']]
                for w in window_dict_zone:
                    object_field = 'Program_Line_' + str(k)
                    new_object[object_field] = 'SET Control_Win_' + str(w) + value_in[21:]
                    k += 1
            else:
                object_field = 'Program_Line_' + str(k)
                new_object[object_field] = value_in
                k += 1
        i += 1
        if not more_loops_i:
            break
    return idf


def write_to_excel_replace(ws, xlsx_mmv, occupation, surfaces_f_dict, surfaces_nf_dict):
    """
    Adds MMV-specific information (infiltration values) to the energy standard sheet of the xlsx_replace file
    :param ws: The worksheet of the replace_mmv.xlsx file
    :param xlsx_mmv: The .xlsx file with MMV procedure instructions
    :param occupation: the occupation of the chosen archetype (e.g. SFH)
    :param surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :param surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    """
    df = load_xlsx_data(xlsx_mmv, 'AFN')
    last_row = ws.max_row
    c = 1
    energy_standards = ['ZEB', 'efficient', 'standard', 'non-standard']
    variables = ['coefficient', 'exponent']
    for k, v in surfaces_f_dict.items():  # multiplication for coefficients is NOT needed
        surface_group = v['Surface_Group']
        for var in variables:
            for en_st in energy_standards:
                idfobject = 'AirflowNetwork:MultiZone:Component:DetailedOpening'
                name = 'Opening_' + str(k)
                if var == 'coefficient':
                    objectfield = 'Air_Mass_Flow_Coefficient_When_Opening_is_Closed'
                else:
                    objectfield = 'Air_Mass_Flow_Exponent_When_Opening_is_Closed'
                value = find_value_in_AFN_df(df, surface_group, en_st, var)
                ws = fill_excel_replace_row(ws, last_row + c, occupation, en_st, idfobject, name, objectfield, value)
                c += 1
    for k, v in surfaces_nf_dict.items():  # multiplication for coefficients is needed, units 'kg/s.m2' -> 'kg/s'
        surface_group = v['Surface_Group']
        for var in variables:
            for en_st in energy_standards:
                idfobject = 'AirflowNetwork:MultiZone:Surface:Crack'
                name = 'Crack_' + str(k)
                if var == 'coefficient':
                    objectfield = 'Air_Mass_Flow_Coefficient_at_Reference_Conditions'
                    value = find_value_in_AFN_df(df, surface_group, en_st, var) * float(v['Area'])
                else:
                    objectfield = 'Air_Mass_Flow_Exponent'
                    value = find_value_in_AFN_df(df, surface_group, en_st, var)
                ws = fill_excel_replace_row(ws, last_row + c, occupation, en_st, idfobject, name, objectfield, value)
                c += 1

    return


def fill_excel_replace_row(ws, row, occupation, energy_standard, idfobject, name, objectfield, value):
    """
    Fills one row of the energy standard sheet of the xlsx_replace file
    :param ws: Worksheet
    :param row: Row number
    :param occupation: the occupation of the chosen archetype (e.g. SFH)
    :param energy_standard: Energy standard name
    :param idfobject: Idf object type
    :param name: Idf object name
    :param objectfield: Idf object field
    :param value: Value for a given idf object field
    :return ws: Worksheet with new row
    """
    ws.cell(column=1, row=row, value=occupation)
    ws.cell(column=2, row=row, value=energy_standard)
    ws.cell(column=3, row=row, value=idfobject)
    ws.cell(column=4, row=row, value=name)
    ws.cell(column=5, row=row, value=objectfield)
    ws.cell(column=6, row=row, value=value)
    return ws


def find_value_in_AFN_df(df, surface_group, en_st, var):
    """
    Returns a value from a dataframe for given conditions of surface group, energy standard, and variable type
    :param df: Pandas dataframe with AirflowNetwork (AFN) data
    :param surface_group: Surface group (e.g., Walls internal)
    :param en_st: Energy standard (e.g., standard, ZEB)
    :param var: Variable type (exponent or coefficient)
    :return value: Value of the variable for the given conditions
    """
    rslt_df = df.loc[(df['surface group'] == surface_group) &
                     (df['energy standard BuildME'] == en_st) &
                     (df['variable'] == var)]
    value = rslt_df['value'].iloc[0]
    return value


def estimate_no_of_floors(idf, zone_dict_mmv):
    """
    Looks for the highest zone within a building to estimate the total number of floors
    :param idf: The .idf file
    :param zone_dict_mmv: Dictionary of zones suitable for MMV (their names and the people object that belongs to it)
    :return no_of_floors: Number of floors
    """
    mmv_zone_list = [i['Zone_Name'] for i in zone_dict_mmv.values()]
    ceiling_heights = []
    for obj in idf.idfobjects['BuildingSurface:Detailed']:
        if obj.Zone_Name in mmv_zone_list:
            ceiling_height = max(abs(pt[2]) for pt in obj.coords)
            zone_obj = [z for z in idf.idfobjects['Zone'] if z.Name == obj.Zone_Name][0]
            multiplier = zone_obj.Multiplier
            if multiplier == "":
                multiplier = 1
            else:
                multiplier = float(zone_obj.Multiplier)
            ceiling_height = ceiling_height * multiplier + zone_obj.Z_Origin
            ceiling_heights.append(ceiling_height)
    max_height = max(ceiling_heights)
    no_of_floors = round(max_height / 2.6)  # assuming that the height of one floor is 2.6 m on average
    return no_of_floors


def create_or_update_excel_replace(occupation, xlsx_mmv, surfaces_f_dict, surfaces_nf_dict, dir_replace_mmv):
    """
    Checks if the excel replace_mmv.xlsx file exists, creates/opens it to write data about the chosen archetype,
    the data is necessary to adjust the archetype to different energy standards
    :param occupation: the occupation of the chosen archetype (e.g. SFH)
    :param xlsx_mmv: The .xlsx file with MMV procedure instructions
    :param surfaces_f_dict: Dictionary of fenestration surfaces (area, surface group, list of surfaces' names)
    :param surfaces_nf_dict: Dictionary of non-fenestration surfaces (area, surface group, list of surfaces' names)
    """
    if os.path.exists(dir_replace_mmv):
        wb = openpyxl.load_workbook(filename=dir_replace_mmv)
        ws_info = wb['info']
        ws_info.cell(column=2, row=2, value=datetime.datetime.now().strftime("%b-%d-%Y, %H:%M"))
        ws = wb['en-standard']
    else:
        wb, ws = create_empty_replace_mmv()
    print("Updating replace_mmv.xlsx file")
    write_to_excel_replace(ws, xlsx_mmv, occupation, surfaces_f_dict, surfaces_nf_dict)
    wb.save(filename=dir_replace_mmv)
    wb.close()
    return


def create_empty_replace_mmv():
    """
    Creates an empty replace_mmv.xlsx file. It contains no data, just metadata and column names
    """
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "info"
    ws_info.cell(column=1, row=1, value="Created:")
    ws_info.cell(column=1, row=2, value="Modified:")
    ws_info.cell(column=2, row=1, value=datetime.datetime.now().strftime("%b-%d-%Y, %H:%M"))
    ws_info.cell(column=2, row=2, value=datetime.datetime.now().strftime("%b-%d-%Y, %H:%M"))
    ws = wb.create_sheet('en-standard')
    ws.cell(column=1, row=1, value="Occupation")
    ws.cell(column=2, row=1, value="standard")
    ws.cell(column=3, row=1, value="idfobject")
    ws.cell(column=4, row=1, value="Name")
    ws.cell(column=5, row=1, value="objectfield")
    ws.cell(column=6, row=1, value="Value")
    return wb, ws
