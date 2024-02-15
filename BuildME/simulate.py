"""
Functions to perform the actual simulations.

Copyright: Niko Heeren, 2019
"""
import datetime
import multiprocessing as mp
import os
import shutil
from time import sleep
import pandas as pd
from tqdm import tqdm
from eppy.modeleditor import IDF
import openpyxl
import numpy as np
from BuildME import energy, material, settings, batch, mmv


def validate_ep_version(idf_files, crash=True):
    """
    Walks through all archetype files and verify matching energyplus version
    :param idf_files: List of idf files to check
    :param crash: Raise Error if true and an error is found. Else script will continue.
    """
    # Check if energyplus version matches the one of the binary
    idd = os.path.abspath(os.path.join(settings.ep_path, "Energy+.idd"))
    with open(idd, mode='r') as f:
        bin_ver = f.readline().strip().split()[1]  # Extract the version from the IDD file's first line, e.g. '!IDD_Version 9.2.0'

    if bin_ver != settings.ep_version:
        err = "WARNING: energyplus version in settings (%s) does not match implied version (%s) from path (%s)" \
              % (settings.ep_version, bin_ver, idd)
        if crash:
            raise AssertionError(err)
        else:
            print(err)
    # https://stackoverflow.com/a/18394205/2075003

    for idff in idf_files:
        with open(idff) as f:
            for line in f:
                if '!- Version Identifier' in line:
                    if settings.ep_version[:-2] not in line:
                        print("WARNING: '%s' has the wrong energyplus version; %s" % (idff, line))


def create_mmv_variant(idf_path, ep_dir, archetype):
    """
    Create MMV variants for the archetypes unless already created
    :param idf_path: path to the IDF file
    :param ep_dir: EnergyPlus directory
    :param archetype: archetype name
    """
    xlsx_mmv = './data/afn-mmv-implementation.xlsx'
    idf_path_original = idf_path.replace('_auto-MMV.idf', '.idf')
    idf_f = read_idf(ep_dir, idf_path_original)
    dictionaries = mmv.create_dictionaries(idf_f, archetype)
    #TODO: if the archetype doesn't have the proper AFN objects, print a message and stop execution 
    flag = mmv.check_if_mmv_zones(dictionaries)
    if flag:  # if the archetype can be created
        print(f"Creating the MMV variant for %s..." % os.path.relpath(idf_path, settings.archetypes))
        idf_mmv = mmv.change_archetype_to_MMV(idf_f, dictionaries, xlsx_mmv)
        if os.path.isfile(idf_path) is True:
            os.remove(idf_path)
        idf_mmv.saveas(idf_path)
    else:
        raise Exception(f"The MMV variant for {archetype} cannot be created")
    return


def find_cpus(method=settings.cpus):
    """
    Returns the number of CPUs available on the system. This can then be used in calculate_energy_mp for example.
     Not using the maximum number of CPUs is sometimes beneficial, because the computer can shuffle around data, etc.
     more efficiently.

    :param method: Method to determine. 'max' will return the maximum number. 'auto' will (hopefully) return the best
                    number.
    """
    if isinstance(method, int):
        return method
    available_cpus = mp.cpu_count()
    if method == 'max':
        return available_cpus
    elif method == 'auto':
        if available_cpus <= 4:
            return available_cpus
        elif available_cpus > 4:
            return available_cpus - 1
    else:
        raise AssertionError("Method '%s' unknown." % method)


def read_idf(ep_dir, idf_path):
    """
    Reads an idf file using eppy
    :param ep_dir: EnergyPlus directory
    :param idf_path: path to the idf file
    :returns: idf file
        """
    idd = os.path.abspath(os.path.join(ep_dir, "Energy+.idd"))
    IDF.setiddname(idd)
    with open(idf_path, 'r') as infile:
        idf = IDF(infile)
    return idf


def apply_obj_name_change(idf, aspect, aspect_value):
    """
    Searches for a defined string (replace_str) in defined IDF fiel objects and prelaces them with a replacement
      string (replacer).
    :param idf: IDF file
    :param aspect: Name of the aspect to be replaced, e.g., 'en_std'
    :param aspect_value: String to search and replace, e.g. 'ZEB'
    :returns: Modified idf file
    """
    objects = ['FenestrationSurface:Detailed', 'BuildingSurface:Detailed', 'Door', 'Window', 'InternalMass']
    replace_str = '-'+aspect+'-replaceme'
    flag_replaceme = False
    for obj_type in objects:
        for obj in idf.idfobjects[obj_type.upper()]:
            if replace_str in obj.Construction_Name:
                # replace the item
                obj.Construction_Name = obj.Construction_Name.replace(replace_str, '-' + aspect_value)
                flag_replaceme = True
    if not flag_replaceme:
        print(f"Warning: No replaceme strings found for aspect {aspect} in building "
              f"{idf.idfobjects['BUILDING'][0].Name}")
    return idf


def apply_rule_from_excel(idf_f, aspect, aspect_value, archetype, csv_folder):
    """
    The function will use the values in the replacement csv spreadsheet and replace them in the idf file.
    :param idf_f: Un-modified / original idf file
    :param aspect: res or en-std
    :param aspect_value:
    :param archetype: archetype name
    :param csv_folder: the location of the spreadsheet with replacement rules
    :return: Modified idf file
    """
    csv_dir = os.path.join(csv_folder, "replace_"+aspect+".csv")
    csv_replace = pd.read_csv(csv_dir, index_col=[0, 1])
    csv_replace.sort_index(inplace=True)
    try:
        csv_data = csv_replace.loc(axis=0)[archetype, aspect_value]  # if no replacement, the variable is empty
    except KeyError:
        print(f"WARNING: Did not find any replacement for archetype {archetype} and {aspect} = {aspect_value} "
              f"in {csv_dir}")
    else:
        if type(csv_data) is pd.Series:
            csv_data = csv_data.to_frame().T
        for row in csv_data.iterrows():
            if row[1]['Value'] == 'skip':
                continue
            obj = idf_f.getobject(row[1]['idfobject'].upper(), row[1]['Name'])
            if obj is not None:
                obj[row[1]['objectfield']] = row[1]['Value']
            else:
                print(f"WARNING: Did not find the EnergyPlus object {(row[1]['idfobject'].upper(), row[1]['Name'])} "
                      f"in the IDF of archetype {archetype}")
    return idf_f


def copy_idf_file(idf_path, out_dir, replace_dict, archetype, ep_dir, replace_csv_dir):
    """
    Copies the chosen idf file to the building simulation folder
    :param idf_path: path to the IDF file
    :param out_dir: output folder directory
    :param replace_dict: dictionary with BuildME replacement aspects
    :param archetype: archetype name
    :param ep_dir: EnergyPlus directory
    :param replace_csv_dir: folder with replacement csv files, e.g., 'replace-en-std.csv'
    """
    idf_path_new = os.path.join(out_dir, 'in.idf')
    shutil.copy2(idf_path, idf_path_new)
    idf_file = read_idf(ep_dir, idf_path_new)
    if replace_dict:
        for aspect, aspect_value in replace_dict.items():
            idf_file = apply_obj_name_change(idf_file, aspect, aspect_value)
            if replace_csv_dir is not None and os.path.exists(replace_csv_dir):
                idf_file = apply_rule_from_excel(idf_file, aspect, aspect_value, archetype, replace_csv_dir)
    if archetype in os.path.basename(out_dir):
        idf_file.idfobjects['Building'.upper()][0].Name = os.path.basename(out_dir)
    idf_file.saveas(idf_path_new)
    return


def calculate_energy(batch_sim=None, idf_path=None, out_dir=None, ep_dir=None, archetype=None, replace_dict=None,
                     parallel=False, clear_folder=False, last_run=False, replace_csv_dir=None, epw_path=None,
                     keep_all=False):
    """
    Initiates the calculation of energy demand
    :param batch_sim: dictionary with batch simulation information
    :param idf_path: path to the IDF file
    :param out_dir: output folder directory
    :param ep_dir: EnergyPlus directory
    :param archetype: archetype name
    :param replace_dict: dictionary with BuildME replacement aspects
    :param parallel: True if parallel simulations (multiprocessing) should be performed (default: False)
    :param clear_folder: True if the simulation folder should be cleared before the simulation (default: False)
    :param last_run: True if the last simulation run should be loaded (default: False)
    :param replace_csv_dir: folder with replacement csv files, e.g., 'replace-en-std.csv'
    :param epw_path: path to the EPW file with weather data
    :param keep_all: boolean indicating whether to keep all simulation files (incl. the .eso file)
    """
    print("Initiating energy demand simulation...")
    # check if all necessary variables are defined
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if batch_sim is None:  # for a standalone simulation
        check_input_variables_standalone(ep_dir, idf_path, out_dir, replace_csv_dir, clear_folder)
        if epw_path is None or not os.path.exists(epw_path):
            print(f"Weather file (defined as {epw_path}) was not not found. "
                  f"\nA dummy weather file for New York city (US) will be used instead.")
            epw_path = os.path.join(settings.climate_files_path, 'USA_NY_New.York-dummy.epw')
        if not os.path.exists(os.path.join(out_dir, 'in.idf')):
            copy_idf_file(idf_path, out_dir, replace_dict, archetype, ep_dir, replace_csv_dir)
        validate_ep_version([os.path.join(out_dir, 'in.idf')])
        # perform actual simulation
        energy.perform_energy_calculation(out_dir, ep_dir, epw_path, keep_all)
    else:  # for batch simulation
        ep_dir = settings.ep_path
        replace_csv_dir = settings.replace_csv_dir
        # copy the necessary files
        for sim in batch_sim:
            idf_path = batch_sim[sim]['archetype_file']
            out_dir = batch_sim[sim]['run_folder']
            epw_path = batch_sim[sim]['climate_file']
            archetype = batch_sim[sim]['occupation']
            replace_dict = batch_sim[sim]['replace_dict']
            if not os.path.exists(os.path.join(out_dir, 'in.idf')): # if the in.idf hasn't been created yet
                if not os.path.exists(idf_path) and batch_sim[sim]['cooling'] == 'MMV':
                    create_mmv_variant(idf_path, ep_dir, archetype)
                copy_idf_file(idf_path, out_dir, replace_dict, archetype, ep_dir, replace_csv_dir)
        validate_ep_version(list(set([batch_sim[sim]['archetype_file'] for sim in batch_sim])))  # list with no duplicates
        
        # perform the simulation (ordinary or parallel)
        if parallel is False:  # ordinary simulation
            for sim in tqdm(batch_sim):
                out_dir = batch_sim[sim]['run_folder']
                epw_path = batch_sim[sim]['climate_file']
                if not os.path.exists(epw_path):
                    print(f"Weather file (defined as {epw_path}) was not not found. "
                          f"\nA dummy weather file for New York city (US) will be used instead.")
                    epw_path = os.path.join(settings.climate_files_path, 'USA_NY_New.York-dummy.epw')
                # perform actual simulation
                energy.perform_energy_calculation(out_dir, ep_dir, epw_path, keep_all)
        else: # parallel simulation
            cpus = find_cpus()
            print("Perform energy simulation on %s CPUs..." % cpus)
            pool = mp.Pool(processes=cpus)
            m = mp.Manager()
            q = m.Queue()
            pbar = tqdm(total=len(batch_sim), smoothing=0.1, unit='sim')
            args = [(batch_sim[sim]['run_folder'], ep_dir, batch_sim[sim]['climate_file'], keep_all, q, i)
                    for i, sim in enumerate(batch_sim)]
            result = pool.map_async(energy.perform_energy_calculation_mp, args)
            old_q = 0
            while not result.ready():
                if q.qsize() > old_q:
                    pbar.update(q.qsize() - old_q)
                else:
                    pbar.update(0)
                old_q = q.qsize()
                sleep(0.2)
            pool.close()
            pool.join()
            pbar.close()
    print('Energy demand simulation finished.')
    return


def calculate_materials(batch_sim=None, idf_path=None, out_dir=None, ep_dir=None, archetype=None, replace_dict=None,
                        clear_folder=False, last_run=False, replace_csv_dir=None, atypical_materials=None,
                        ifsurrogates=True, surrogates=None, region=None):
    """
    Initiates the calculation of material demand
    :param batch_sim: dictionary with batch simulation information
    :param idf_path: path to the IDF file
    :param out_dir: output folder directory
    :param ep_dir: EnergyPlus directory
    :param archetype: archetype name
    :param replace_dict: dictionary with BuildME replacement aspects
    :param clear_folder: True if the simulation folder should be cleared before the simulation (default: False)
    :param last_run: True if the last simulation run should be loaded (default: False)
    :param replace_csv_dir: folder with replacement csv files, e.g., 'replace-en-std.csv'
    :param atypical_materials: pandas dataframe with thicknesses and densities of atypical materials (also accepts dict)
    :param ifsurrogates: True if surrogate calculations are requested (default: False)
    :param surrogates: dictionary with surrogate element information (pandas dataframe is also accepted)
    :param region: name of the region (only necessary when surrogates is a pandas dataframe with a "region" column)
    """
    print("Initiating material demand simulation...")
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if batch_sim is None:  # for a standalone simulation
        check_input_variables_standalone(ep_dir, idf_path, out_dir, replace_csv_dir, clear_folder)
        if not os.path.exists(os.path.join(out_dir, 'in.idf')):
            copy_idf_file(idf_path, out_dir, replace_dict, archetype, ep_dir, replace_csv_dir)
        if ifsurrogates:
            if surrogates is None:
                raise Exception("Surrogate element calculations requested but no surrogate dictionary given")
            elif type(surrogates) == pd.core.frame.DataFrame:
                surrogates = convert_surrogates_df_to_dict(surrogates, archetype, replace_dict, region)
        validate_ep_version([os.path.join(out_dir, 'in.idf')])
        # perform actual simulation
        idf_file = read_idf(ep_dir, os.path.join(out_dir, 'in.idf'))
        atypical_materials = check_atypical_materials(idf_file, atypical_materials, out_dir, config=False)
        material.perform_materials_calculation(idf_file, out_dir, atypical_materials, surrogates,
                                               ifsurrogates, replace_dict=replace_dict)
    else:  # for a batch simulation
        ep_dir = settings.ep_path
        replace_csv_dir = settings.replace_csv_dir
        atypical_materials = settings.atypical_materials
        # copy the necessary files
        for sim in batch_sim:
            idf_path = batch_sim[sim]['archetype_file']
            out_dir = batch_sim[sim]['run_folder']
            archetype = batch_sim[sim]['occupation']
            replace_dict = batch_sim[sim]['replace_dict']
            if not os.path.exists(os.path.join(out_dir, 'in.idf')):
                if not os.path.exists(idf_path) and batch_sim[sim]['cooling'] == 'MMV':
                    create_mmv_variant(idf_path, ep_dir, archetype)
                copy_idf_file(idf_path, out_dir, replace_dict, archetype, ep_dir, replace_csv_dir)
        validate_ep_version(list(set([batch_sim[sim]['archetype_file'] for sim in batch_sim])))  # list with no duplicates
        # perform actual simulation
        for sim in tqdm(batch_sim):
            out_dir = batch_sim[sim]['run_folder']
            archetype = batch_sim[sim]['occupation']
            replace_dict = batch_sim[sim]['replace_dict']
            region = batch_sim[sim]['climate_region']
            if ifsurrogates:
                surrogates = convert_surrogates_df_to_dict(settings.surrogate_elements, archetype, replace_dict, region)
            idf_file = read_idf(ep_dir, os.path.join(out_dir, 'in.idf'))
            atypical_materials = check_atypical_materials(idf_file, atypical_materials, out_dir, config=True)
            material.perform_materials_calculation(idf_file, out_dir, atypical_materials, surrogates,
                                                   ifsurrogates, replace_dict=replace_dict)
    print('Material demand simulation finished.')
    return


def check_input_variables_standalone(ep_dir, idf_path, out_dir, replace_csv_dir, clear_folder):
    """
    Checks whether input variables required for a standalone material or energy demand simulation are available
    :param ep_dir: EnergyPlus directory
    :param idf_path: path to the IDF file
    :param out_dir: output folder directory
    :param replace_csv_dir: folder with replacement csv files, e.g., 'replace-en-std.csv'
    :param clear_folder: True if the simulation folder should be cleared before the simulation (default: False)
    """
    if ep_dir is None:
        raise Exception('Energy plus path not given')
    elif not os.path.exists(ep_dir):
        raise Exception(f'Energy plus path does not exist ({ep_dir})')
    if idf_path is None:
        raise Exception('Building archetype file not given')
    elif not os.path.exists(idf_path):
        raise Exception(f'Building archetype file does not exist ({idf_path})')
    if out_dir is None:
        raise Exception('Output folder not given')
    else:
        if os.path.exists(out_dir) and clear_folder is True:
            print(f'Clearing the content of the folder {out_dir}')
            shutil.rmtree(out_dir)
            os.makedirs(out_dir)
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
    if replace_csv_dir is None or not os.path.exists(replace_csv_dir):
        print('The directory with "replace" csv files not provided or the provided one does not exist. '
              'These replace rules will be skipped.')
    return


def convert_surrogates_df_to_dict(surrogates, archetype, replace_dict, region):
    """
    Filter the "surrogates" pandas dataframe according to archetype, region and replace_dict
    and then convert the dataframe into a dict
    :param surrogates: pandas dataframe with surrogate element information
    :param archetype: archetype name
    :param replace_dict: dictionary with BuildME replacement aspects
    :param region: name of the region (only necessary when surrogates is a pandas dataframe with a "region" column)
    """
    df_sur_sim = surrogates[surrogates['occupation'] == archetype].drop(columns='occupation')
    # filtering columns 'region', 'en-std' and 'res' (if they exist)
    # keeps a surrogate dataframe's row if the field matches the given value or is empty (== 0)
    if region is not None:
        if 'region' in df_sur_sim.columns:
            df_sur_sim = df_sur_sim[(df_sur_sim['region'] == region) | (df_sur_sim['region'] == 0)]
            df_sur_sim = df_sur_sim.drop(columns='region')
    else:
        if 'region' in df_sur_sim.columns:
            df_sur_sim = df_sur_sim.drop(columns='region')
    if replace_dict is not None:
        for k, v in replace_dict.items():
            if k in df_sur_sim.columns:
                df_sur_sim = df_sur_sim[(df_sur_sim[k] == v) | (df_sur_sim[k] == 0)]
                df_sur_sim = df_sur_sim.drop(columns=k)
    else:
        for k in ['en-std', 'res']:  # if replace_dict is None, drop the columns with en-std and res
            if k in df_sur_sim.columns:
                df_sur_sim = df_sur_sim.drop(columns=k)
    surrogate_duplicates = df_sur_sim['surrogate'].duplicated()
    if surrogate_duplicates.any():
        df_sur_sim = df_sur_sim[~surrogate_duplicates]  # keep the first entry
        print('Warning: duplicates found in the "surrogates" dataframe. Only the first entry will be kept.')
    df_sur_sim = df_sur_sim.set_index('surrogate')
    surrogates_dict = df_sur_sim.to_dict('index')
    if not surrogates_dict:
        print(f'Warning: No surrogate elements found for archetype {archetype} '
              f'and aspects {replace_dict}. '
              f'\n\t Surrogate element calculations will be skipped.')
    return surrogates_dict


def check_atypical_materials(idf_file, atypical_materials, out_dir, config=True):
    """
    Check if all atypical materials (with no density and/or thickness data) have externally defined data
    :param idf_file: IDF file
    :param atypical_materials: pandas dataframe with thicknesses and densities of atypical materials (also accepts dict)
    :param out_dir: output folder directory
    :param config: True if the configuration file should be used (to add the missing materials into the file)
    """
    if atypical_materials is None:
        if os.path.exists(os.path.join(out_dir, 'atypical_materials.csv')):
            print('Atypical materials automatically read from "atypical_materials.csv".')
            atypical_materials = pd.read_csv(os.path.join(out_dir, 'atypical_materials.csv'), index_col=0)
        else:
            atypical_materials = pd.DataFrame()
    elif type(atypical_materials) is dict:
        density = [v for dic in atypical_materials.values() for k, v in dic.items() if 'density' in k]
        thickness = [v for dic in atypical_materials.values() for k, v in dic.items() if 'thickness' in k]
        atypical_materials = pd.DataFrame({'density (kg/m3)': density, 'thickness (m)': thickness},
                                          index=atypical_materials.keys())
    obj_types = ['Material:NoMass', 'Material:InfraredTransparent', 'Material:AirGap',
                 'Material:RoofVegetation', 'WindowMaterial:SimpleGlazingSystem', 'WindowMaterial:Glazing',
                 'WindowMaterial:GlazingGroup:Thermochromic', 'WindowMaterial:Glazing:RefractionExtinctionMethod',
                 'WindowMaterial:Gas', 'WindowMaterial:GasMixture', 'WindowMaterial:Gap', 'WindowMaterial:Shade',
                 'WindowMaterial:ComplexShade', 'WindowMaterial:Blind', 'WindowMaterial:Screen']
    obj_types_with_thickness = ['Material:RoofVegetation', 'WindowMaterial:Glazing',
                                'WindowMaterial:Glazing:RefractionExtinctionMethod', 'WindowMaterial:Gas',
                                'WindowMaterial:GasMixture', 'WindowMaterial:Gap', 'WindowMaterial:Shade']
    weird_mats = [obj for obj_type in obj_types for obj in idf_file.idfobjects[obj_type.upper()]]
    unknown_materials = {}
    for mat in weird_mats:
        if mat.Name not in atypical_materials.index:
            unknown_materials[mat.Name] = mat.obj[0]
    if unknown_materials:
        if config is False:
            thickness_list = ['?' if mat not in obj_types_with_thickness
                              else 'defined in ep' for mat in unknown_materials.values()]
            df = pd.DataFrame({'density (kg/m3)': ['?'], 'thickness (m)': thickness_list}, index=unknown_materials.keys())
            if os.path.exists(os.path.join(out_dir, 'atypical_materials.csv')):
                df_old = pd.read_csv(os.path.join(out_dir, 'atypical_materials.csv'), index_col=0)
                df = pd.concat([df_old, df])
            df.to_csv(os.path.join(out_dir, 'atypical_materials.csv'))
            raise Exception(f'The following materials were not found in the atypical materials dictionary: '
                            f'{list(unknown_materials.keys())}.'
                            f'\nThese materials were added to file "atypical_materials.csv".')
        else:
            wb = openpyxl.load_workbook(filename=settings.config_file)
            ws = wb['atypical materials']
            last_row = ws.max_row
            c = 1
            for mat, mat_type in unknown_materials.items():
                ws.cell(column=3, row=last_row + c, value=mat)
                ws.cell(column=4, row=last_row + c, value='?')
                if mat_type in obj_types_with_thickness:
                    ws.cell(column=5, row=last_row + c, value='defined in ep')
                else:
                    ws.cell(column=5, row=last_row + c, value='?')
                c += 1
            wb.save(filename=settings.config_file)
            wb.close()
            raise Exception(f'The following materials were not found in the atypical materials dictionary: '
                            f'\n\t {list(unknown_materials.keys())}.'
                            f"\n\t These materials were added in sheet 'atypical materials' of the file "
                            f'{os.path.basename(settings.config_file)}')
    unspecified_materials = atypical_materials[(atypical_materials['density (kg/m3)'] == '?')
                                               | (atypical_materials['thickness (m)'] == '?')]
    if list(unspecified_materials.index):
        raise Exception(f"The atypical materials dictionary includes entries with '?' instead of values "
                        f"for the following materials: \n\t {list(unspecified_materials.index)}.")
    return atypical_materials


def aggregate_energy(batch_sim=None, last_run=False, folders=None, unit='MJ'):
    """
    Reads the EnergyPlus result file 'eplusout.csv' and aggregates the results (sums the column)
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded (default: False)
    :param folders: a list of directories (required only when batch_sim is None)
    :param unit: energy units in the output file - kWh, J or MJ (default)
    :returns: df_results
    """
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if batch_sim is None:
        if folders is None:
            raise Exception('Folders not given')
    else:
        folders = [batch_sim[sim]['run_folder'] for sim in batch_sim]
    units = ['J', 'MJ', 'kWh']
    if unit is None:
        unit = 'MJ'
    elif unit not in units:
        print(f'Unit {unit} not recognized. MJ will be used instead')
        unit = 'MJ'
    multipliers = [1, 1/10**6, 1/(3.6*10**6)]
    multiplier = multipliers[units.index(unit)]
    # Note the trailing whitespace at the end of "InteriorEquipment:Electricity [J](Hourly) "
    results_to_collect = ("Heating:EnergyTransfer [J](Hourly)",	"Cooling:EnergyTransfer [J](Hourly)",
                          "InteriorLights:Electricity [J](Hourly)", "InteriorEquipment:Electricity [J](Hourly) ")
    for folder in folders:
        ep_file = os.path.join(folder, 'eplusout.csv')
        ep_out = pd.read_csv(ep_file)
        df_results = ep_out.loc[:, results_to_collect].sum()*multiplier
        df_results.index = [i.split(' [')[0] for i in df_results.index]
        df_results = df_results.reset_index()
        cols = list(df_results.columns)
        new_cols = ['EnergyPlus output variable', 'Value']
        df_results = df_results.rename(columns={k: new_cols[i] for i, k in enumerate(cols)})
        df_results['Unit'] = unit
        df_results = df_results[['EnergyPlus output variable', 'Unit', 'Value']]
        total = pd.DataFrame([['TOTAL', unit, df_results['Value'].sum()]], columns=df_results.columns)
        df_results = pd.concat([df_results, total], ignore_index=True)
        df_results.to_csv(os.path.join(folder, 'energy_demand.csv'), index=False)
    return df_results


def aggregate_materials(batch_sim=None, last_run=False, aggregation_categories=None, folders=None):
    """
    Aggregate material types into categories, e.g., concrete, cement, wood and wood products
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded (default: False)
    :param aggregation_categories: dict with materials and their aggregation categories (also accepts pandas dataframe)
    :param folders: a list of directories (required only when batch_sim is None)
    """
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if batch_sim is None:
        if folders is None:
            raise Exception('Folders not given')
        if aggregation_categories is None:
            if len(folders) == 1:
                out_dir = folders[0]
            else:
                out_dir = os.path.dirname(folders[0])
            if os.path.exists(os.path.join(out_dir, 'aggregation_categories.csv')):
                print(f'Aggregation categories automatically read from "aggregation_categories.csv" placed in {out_dir}.')
                aggregation_categories = pd.read_csv(os.path.join(out_dir, 'aggregation_categories.csv'), index_col=0)
                aggregation_categories = aggregation_categories.T.to_dict(orient='list')
                aggregation_categories = {k: v[0] for k, v in aggregation_categories.items()}
            else:
                aggregation_categories = {}
        elif type(aggregation_categories) == pd.core.frame.DataFrame:
            aggregation_categories = aggregation_categories.T.to_dict(orient='list')
            aggregation_categories = {k: v[0] for k, v in aggregation_categories.items()}
    else:
        if aggregation_categories is None:
            aggregation_categories = settings.material_aggregation
        folders = [batch_sim[sim]['run_folder'] for sim in batch_sim]
    unknown_materials = []
    for folder in folders:
        df = pd.read_csv(os.path.join(folder, 'mat_demand.csv'))
        mapping = df['Material name'].map(aggregation_categories)
        df['Material type'] = mapping
        unknown_materials = unknown_materials + df[df['Material type'].isna()]['Material name'].values.tolist()
        df['Material type'] = df['Material type'].replace(np.nan, '?')
        df = df[['Material name', 'Material type', 'Unit', 'Value']]
        filename = os.path.join(folder, 'mat_demand_categorized.csv')
        df.to_csv(filename, index=False)
        df = df.drop(columns='Material name')
        df = df.groupby(['Material type', 'Unit']).sum()
        df = df.reset_index()
        total = pd.DataFrame([['TOTAL', 'kg', df['Value'].sum()]], columns=df.columns)
        df = pd.concat([df, total], ignore_index=True)
        filename = os.path.join(folder, 'mat_demand_aggregated.csv')
        df.to_csv(filename, index=False)
    unknown_materials = list(set(unknown_materials))  # deleting duplicates
    if unknown_materials:
        if batch_sim is None:
            df = pd.DataFrame({'material category': ['?']}, index=unknown_materials)
            df.to_csv(os.path.join(out_dir, 'aggregation_categories.csv'))
            print(f'The following materials were not found in the material aggregation dictionary: '
                  f'\n\t{unknown_materials}'
                  f'\n\tThese materials were added to file "aggregation_categories.csv" placed in {out_dir}.'
                  f"\n\tUnless the aggregation category is specified, "
                  f" these materials will continue to be classified as '?'.")
        else:
            wb = openpyxl.load_workbook(filename=settings.config_file)
            ws = wb['material aggregation']
            last_row = ws.max_row
            c = 1
            for i in unknown_materials:
                ws.cell(column=3, row=last_row + c, value=i)
                ws.cell(column=4, row=last_row + c, value='?')
                c += 1
            wb.save(filename=settings.config_file)
            wb.close()
            print(f'The following materials were not found in the material aggregation dictionary: '
                  f'\n\t{unknown_materials}'
                  f"\n\tThese materials were added in sheet 'material aggregation' of the file"
                  f"{os.path.basename(settings.config_file)}. "
                  f"\n\tUnless the aggregation category is specified,"
                  f" these materials will continue to be classified as '?'.")
    return


def calculate_intensities(batch_sim=None, last_run=False, results=None, folders=None, ref_area='total_floor_area'):
    """
    Calculates intensities of the results (energy or material demand per square meter of floor area)
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded
    :param results: the names of csv files with energy and material results
    :param folders: a list of directories (required only when batch_sim is None)
    :param ref_area: reference: 'total_floor_area' (default), 'floor_area_occupied' or 'floor_area_conditioned'
    """
    if results is None:
        results = ['energy_demand.csv', 'mat_demand.csv', 'mat_demand_aggregated.csv']
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if ref_area not in ['floor_area_occupied', 'floor_area_conditioned', 'total_floor_area'] \
            and type(ref_area) not in [int, float]:
        raise Exception(f'Reference area {ref_area} is unknown')
    if batch_sim is None:
        if folders is None:
            raise Exception('Folders not given')
    else:
        folders = [batch_sim[sim]['run_folder'] for sim in batch_sim]
    for folder in folders:
        if ref_area in [int, float]:
            area = ref_area
        else:
            try:
                df_geom = pd.read_csv(os.path.join(folder, 'geom_stats.csv'), index_col='Geometry statistics')
            except FileNotFoundError as e:
                raise Exception('No geometry data available. Please perform material calculations first.') from e
            # df_geom.index = df_geom['Geometry statistics']
            area = df_geom.loc[ref_area].Value
        for name in results:
            new_name = name.replace('.csv', '_m2.csv')
            try:
                df = pd.read_csv(os.path.join(folder, name))
            except FileNotFoundError:
                pass
            else:
                df['Value'] = df['Value']/float(area)
                df['Unit'] = df['Unit']+'/m2'
                df.to_csv(os.path.join(folder, new_name), index=False)
    return


def collect_results(batch_sim=None, last_run=False, results=None, folders=None, combinations=None):
    """
    Collect results from multiple simulations into one summary file
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded
    :param results: the names of csv files with energy and material results
    :param folders: a list of directories (required only when batch_sim is None)
    :param combinations: a dictionary with the selected BuildME aspects and their values
    """
    if results is None:
        results = ['energy_demand.csv', 'geom_stats.csv', 'mat_demand.csv', 'mat_demand_categorized.csv',
                   'mat_demand_aggregated.csv', 'energy_demand_m2.csv', 'mat_demand_aggregated_m2.csv',
                   'mat_demand_m2.csv']
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if combinations is None:
        combinations = settings.debug_combinations
    aspect_names = ['region'] + list(list(combinations.values())[0].keys())
    if batch_sim is None:
        if folders is None:
            raise Exception('Folders not given')
    else:
        folders = [batch_sim[sim]['run_folder'] for sim in batch_sim]
    parent_dir = os.path.dirname(folders[0])
    for name in results:
        summary_name = 'summary_'+name
        summary_df = pd.DataFrame()
        flag = True
        for folder in folders:
            try:
                df = pd.read_csv(os.path.join(folder, name))
            except FileNotFoundError:
                flag = False
                break
            df['Building name'] = os.path.basename(folder)
            summary_df = pd.concat([summary_df, df])
        if flag:
            summary_df.index = pd.MultiIndex.from_tuples(tuple(summary_df['Building name'].str.split('_')))
            summary_df.index.names = aspect_names
            summary_df = summary_df[['Building name']+list(summary_df.columns[:-1])]
            filename = os.path.join(parent_dir, summary_name)
            summary_df.to_csv(filename)
    return


def weighing_climate_region(batch_sim=None, last_run=False, results=None, combinations=None):
    """
    Multiplies each result by its climate region ratio given in aggregate.xlsx.
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded
    :param results: the names of csv files with energy and material results
    :param combinations: a dictionary with the selected BuildME aspects and their values
    """
    if results is None:
        results = ['summary_energy_demand.csv', 'summary_energy_demand_m2.csv']
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if combinations is None:
        combinations = settings.debug_combinations
    aspect_names = ['region'] + list(list(combinations.values())[0].keys())
    parent_dir = os.path.dirname(batch_sim[list(batch_sim.keys())[0]]['run_folder'])
    weights = settings.climate_region_weight
    for name in results:
        try:
            df = pd.read_csv(os.path.join(parent_dir, name))
        except FileNotFoundError:
            pass
        else:
            df_new = pd.merge(df, weights, on=['region', 'climate_region'])
            df_new['Value'] = df_new['Value'] * df_new['share']
            df_new = df_new.drop(columns=['share'])
            df_new['climate_region'] = 'weighed'
            df_new = df_new.groupby(aspect_names).sum()
            filename = os.path.join(parent_dir, name.replace('.csv', '_weighed.csv'))
            df_new.to_csv(filename)
    return


def cleanup(batch_sim=None, last_run=False, archive=True, del_temp=True):
    """
    Convenience function to clean up after successful run.
    :param batch_sim: dictionary with batch simulation information
    :param last_run: True if the last simulation run should be loaded
    :param archive: Zip temporary folders into e.g. '220628-080114.zip'
    :param del_temp: Delete all subfolders
    """
    if last_run:
        batch_sim = batch.find_and_load_last_run()
    if batch_sim is None:
         raise Exception('Variable batch_sim not given')
    parent_dir = os.path.dirname(batch_sim[list(batch_sim.keys())[0]]['run_folder'])
    run = os.path.basename(parent_dir)
    print(run)
    print("Cleaning up...")
    if archive:
        zfile = os.path.join(settings.tmp_path, run)
        if os.path.exists(zfile + '.zip'):
            print("Warning: Archive '%s.zip' already exists" % zfile)
        else:
            print("Compressing temporary folder to '%s.zip'" % zfile)
            shutil.make_archive(zfile, 'zip', os.path.join(settings.tmp_path, run))
    if del_temp:
        rfolders = [batch_sim[d]['run_folder'] for d in batch_sim]
        for f in rfolders:
            shutil.rmtree(f)
        print("Deleted temporary subfolders in '%s/%s/'" % (settings.tmp_path, run))
    print("Cleanup done.")
