"""
Functions to adapt custom IDF files to BuildME.

Use these functions:
to automatically insert replaceme strings on surface objects,
to rename materials/constructions based on standards/resource efficiency scenarios or any other layers(i.e., cohort),
to merge idfs that have different materials/constructions
to update replaceme.yaml or to update replaceme.xlsx
to update material.yaml or to update material.xlsx
to change Output objects' contents in a way that BuildME accepts

Copyright: Sahin AKIN, 2022
"""

import os
from functools import reduce
from io import StringIO
import openpyxl
import yaml
from eppy.modeleditor import IDF

# Make sure that you have selected the correct working directory (BUILDME)
# Standalone Run Requirements
ep_version = '9.2.0'
path_parent = os.path.dirname(os.getcwd())
os.chdir(path_parent)
basepath = os.path.abspath('..')
ep_path = os.path.abspath("..\\bin\\EnergyPlus-9-2-0")
ep_idd = os.path.abspath("..\\bin\\EnergyPlus-9-2-0\\Energy+.idd")
IDF.setiddname(ep_idd)


def create_combined_idflist(folderpath):
    """
    Creates a list containing the generated idf files
    :param folderpath: (in other words, save folder) containing all generated idfs for standards/res scenarios for a single archetype type (e.g. RT only)
    :return: a combined idf list
    """
    print("IDF list is being generated...")
    filepaths = [os.path.join(folderpath, name) for name in os.listdir(folderpath)]
    all_files = []
    for path in filepaths:
        if "-converted" in path:
            with open(path, 'r') as f:
                file = f.read()
                all_files.append(file)

    return all_files


def convert_idf_to_BuildME(idf_path, save_folder_path, replace_string="-en-std-replaceme",
                           replace_string_value="-non-standard", base=False):
    """
    This function creates an individual edited IDF file that is compatible with BuildME framework

    Requirements for a clear run:
    1)Window-Skylight Difference should be addressed in construction name (e.g. name contains "sky")

    Run Order:
    -Adds replaceme fields,
    -Renames Materials,
    -Renames Constructions,
    -Defines non-standard U values; if replace_string_value is set accordingly.

    :param idf_path: path for idf file to be converted, standard file should be used forthe non-standard IDF creation
    :param save_folder_path: New location folder path, where the new IDF will be saved
    :param replace_string: Replaceme content (i.e., -en-std-replaceme-res-replacemem-chrt-replaceme), should start with "-"
    :param replace_string_value: Replacer string corresponding to the replace string (i.e., -standard-RES0-1920), should start with "-"
    :param base: Boolean value to declare whether the IDF file corresponding to a base IDF
    :return: A compatible file with BuildME for an IDF with specific features
    """

    idf1 = IDF(idf_path)
    print("Conversion is initialized, construction and surface objects are being converted...")
    # CONVERTING CONSTRUCTION NAMES AND INSERTING REPLACEME STRINGS:
    # It inserts a replaceme string to the building surface object's construction field, adds a new identical construction to the construction object
    # and renames the construction name corresponding to the replaceme string
    for items in idf1.idfobjects["BuildingSurface:Detailed".upper()]:
        for obj in idf1.idfobjects["Construction".upper()]:
            if obj.Name == items.Construction_Name:

                if items.Surface_Type == "Roof":
                    items.Construction_Name = f"ext-roof{replace_string}"
                    newcons = idf1.copyidfobject(obj)
                    newcons.Name = f"ext-roof{replace_string_value}"

                if items.Surface_Type == "Ceiling":
                    if "int" in items.Construction_Name or "ceiling" in items.Construction_Name:
                        items.Construction_Name = f"int-ceiling{replace_string}"
                        newcons = idf1.copyidfobject(obj)
                        newcons.Name = f"int-ceiling{replace_string_value}"

                if items.Surface_Type == "Floor":
                    if items.Outside_Boundary_Condition == "Surface":
                        items.Construction_Name = f"int-floor{replace_string}"
                        newcons = idf1.copyidfobject(obj)
                        newcons.Name = f"int-floor{replace_string_value}"
                    if items.Outside_Boundary_Condition == "Adiabatic" or items.Outside_Boundary_Condition == "Ground" or items.Outside_Boundary_Condition == "Outdoors" or items.Outside_Boundary_Condition == "GroundSlabPreprocessorAverage":
                        items.Construction_Name = f"ext-floor{replace_string}"
                        newcons = idf1.copyidfobject(obj)
                        newcons.Name = f"ext-floor{replace_string_value}"

                if items.Surface_Type == "Wall":
                    if "int" in items.Construction_Name or "partition" in items.Construction_Name or items.Outside_Boundary_Condition == "Surface":
                        items.Construction_Name = f"int-wall{replace_string}"
                        newcons = idf1.copyidfobject(obj)
                        newcons.Name = f"int-wall{replace_string_value}"
                    else:
                        items.Construction_Name = f"ext-wall{replace_string}"
                        newcons = idf1.copyidfobject(obj)
                        newcons.Name = f"ext-wall{replace_string_value}"

        if 'CONSTRUCTION:FFACTORGROUNDFLOOR' in [x for x in idf1.idfobjects]:
            for floor in idf1.idfobjects["Construction:FfactorGroundFloor".upper()]:
                if floor.Name == items.Construction_Name:
                    if items.Surface_Type == "Floor":
                        if items.Outside_Boundary_Condition == "GroundFCfactorMethod":
                            items.Construction_Name = f"{items.Construction_Name}{replace_string}"
                            newcons = idf1.copyidfobject(floor)
                            newcons.Name = f"{floor.Name}{replace_string_value}"

        if 'CONSTRUCTION:CFACTORUNDERGROUNDWALL' in [x for x in idf1.idfobjects]:
            for wall in idf1.idfobjects["Construction:CfactorUndergroundWall".upper()]:
                if wall.Name == items.Construction_Name:
                    if items.Surface_Type == "Wall":
                        if items.Outside_Boundary_Condition == "GroundFCfactorMethod" or items.Outside_Boundary_Condition == "Adiabatic":
                            items.Construction_Name = f"{items.Construction_Name}{replace_string}"
                            newcons = idf1.copyidfobject(wall)
                            newcons.Name = f"{wall.Name}{replace_string_value}"

    if 'WINDOW' in [x for x in idf1.idfobjects]:
        for fenest in idf1.idfobjects["Window".upper()]:
            fenest.Construction_Name = f"ext-window{replace_string}"

    if 'DOOR' in [x for x in idf1.idfobjects]:
        for fenest in idf1.idfobjects["Door".upper()]:
            fenest.Construction_Name = f"ext-door{replace_string}"

    if 'WINDOWSHADINGCONTROL' in [x for x in idf1.idfobjects]:
        for shade in idf1.idfobjects["WindowShadingControl".upper()]:
            for obj in idf1.idfobjects["Construction".upper()]:
                if shade.Construction_with_Shading_Name == obj.Name:
                    shade.Construction_with_Shading_Name = f"ext-window-wshade{replace_string}"
                    newcons = idf1.copyidfobject(obj)
                    newcons.Name = f"ext-window-wshade{replace_string_value}"
    print("Checking each window,frame...This will take some time!")
    if 'FENESTRATIONSURFACE:DETAILED' in [x for x in idf1.idfobjects]:
        for fenest in idf1.idfobjects["FenestrationSurface:Detailed".upper()]:
            for obj in idf1.idfobjects["Construction".upper()]:
                for frame in idf1.idfobjects["WindowProperty:FrameAndDivider".upper()]:
                    if fenest.Construction_Name == obj.Name:
                        if fenest.Surface_Type == "Window":

                            # there might be skylights:
                            if fenest.Frame_and_Divider_Name == frame.Name:
                                fenest.Frame_and_Divider_Name = f"ext-frame{replace_string}"
                                newcons = idf1.copyidfobject(frame)
                                newcons.Name = f"ext-frame{replace_string_value}"
                                fenest.Construction_Name = f"ext-window{replace_string}"
                                newcons = idf1.copyidfobject(obj)
                                newcons.Name = f"ext-window{replace_string_value}"

                            if "sky" in fenest.Construction_Name:
                                fenest.Construction_Name = f"ext-skywindow{replace_string}"
                                newcons = idf1.copyidfobject(obj)
                                newcons.Name = f"ext-skywindow{replace_string_value}"
                            else:
                                fenest.Construction_Name = f"ext-window{replace_string}"
                                newcons = idf1.copyidfobject(obj)
                                newcons.Name = f"ext-window{replace_string_value}"

                        if fenest.Surface_Type == "GlassDoor":
                            fenest.Construction_Name = f"ext-window{replace_string}"
                            newcons = idf1.copyidfobject(obj)
                            newcons.Name = f"ext-window{replace_string_value}"
                        if fenest.Surface_Type == "Door":
                            fenest.Construction_Name = f"ext-door{replace_string}"
                            newcons = idf1.copyidfobject(obj)
                            newcons.Name = f"ext-door{replace_string_value}"
    print("Still...")
    if 'FENESTRATIONSURFACE:DETAILED' in [x for x in idf1.idfobjects]:
        for fenest in idf1.idfobjects["FenestrationSurface:Detailed".upper()]:
            for obj in idf1.idfobjects["Construction".upper()]:
                for frame in idf1.idfobjects["WindowProperty:FrameAndDivider".upper()]:
                    if fenest.Construction_Name == obj.Name:
                        if fenest.Surface_Type == "Window":

                            # there might be skylights:
                            if fenest.Frame_and_Divider_Name != " ":
                                fenest.Frame_and_Divider_Name = f"ext-frame{replace_string}"
                                newcons = idf1.copyidfobject(frame)
                                newcons.Name = f"ext-frame{replace_string_value}"
                                fenest.Construction_Name = f"ext-window{replace_string}"
                                newcons = idf1.copyidfobject(obj)
                                newcons.Name = f"ext-window{replace_string_value}"

    # Deleting duplicated construction names
    unique = reduce(lambda l, x: l.append(x) or l if x not in l else l, idf1.idfobjects["Construction".upper()], [])
    idf1.removeallidfobjects("CONSTRUCTION")
    for newcons in unique:
        idf1.copyidfobject(newcons)
    if 'CONSTRUCTION:FFACTORGROUNDFLOOR' in [x for x in idf1.idfobjects]:
        unique = reduce(lambda l, x: l.append(x) or l if x not in l else l,
                        idf1.idfobjects["Construction:FfactorGroundFloor".upper()], [])
        idf1.removeallidfobjects("CONSTRUCTION:FFACTORGROUNDFLOOR")
    for newcons in unique:
        idf1.copyidfobject(newcons)
    if 'CONSTRUCTION:CFACTORUNDERGROUNDWALL' in [x for x in idf1.idfobjects]:
        unique = reduce(lambda l, x: l.append(x) or l if x not in l else l,
                        idf1.idfobjects["Construction:CfactorUndergroundWall".upper()], [])
        idf1.removeallidfobjects("CONSTRUCTION:CFACTORUNDERGROUNDWALL")
    for newcons in unique:
        idf1.copyidfobject(newcons)

    if 'WindowProperty:FrameAndDivider'.upper() in [x for x in idf1.idfobjects]:
        unique = reduce(lambda l, x: l.append(x) or l if x not in l else l,
                        idf1.idfobjects["WindowProperty:FrameAndDivider".upper()], [])
        idf1.removeallidfobjects("WindowProperty:FrameAndDivider".upper())
    for newcons in unique:
        idf1.copyidfobject(newcons)

    print("Material objects are being converted...")
    # CONVERTING MATERIALS
    # For the -non-standard version, values are replaced with %30 worse performing numbers based on the standard/base version.
    if "-non-standard" in replace_string_value:
        for items in idf1.idfobjects["Material".upper()]:
            items.Conductivity = float(items.Conductivity) * 1.3
            items.Name = f"{items.Name}{replace_string_value}"
        for items in idf1.idfobjects["Material:NoMass".upper()]:
            items.Thermal_Resistance = float(items.Thermal_Resistance) * 0.7
            items.Name = f"{items.Name}{replace_string_value}"
        for items in idf1.idfobjects["WindowMaterial:SimpleGlazingSystem".upper()]:
            items.UFactor = float(items.UFactor) * 1.3
            items.Name = f"{items.Name}{replace_string_value}"
        if 'CONSTRUCTION:FFACTORGROUNDFLOOR' in [x for x in idf1.idfobjects]:
            for items in idf1.idfobjects["Construction:FfactorGroundFloor".upper()]:
                items.FFactor = float(items.FFactor) * 0.7
        if 'CONSTRUCTION:CFACTORUNDERGROUNDWALL' in [x for x in idf1.idfobjects]:
            for items in idf1.idfobjects["Construction:CfactorUndergroundWall".upper()]:
                items.CFactor = float(items.CFactor) * 1.3

    # Material names are changed as they represent different values across standards, some identifiers are added based on replacer_string.
    else:
        for items in idf1.idfobjects["Material".upper()]:
            items.Name = f"{items.Name}{replace_string_value}"
        for items in idf1.idfobjects["Material:NoMass".upper()]:
            items.Name = f"{items.Name}{replace_string_value}"
        for items in idf1.idfobjects["WindowMaterial:SimpleGlazingSystem".upper()]:
            items.Name = f"{items.Name}{replace_string_value}"
        for items in idf1.idfobjects["WindowMaterial:Shade".upper()]:
            items.Name = f"{items.Name}{replace_string_value}"

    # Construction material layer names are matched with above material changes
    for items in idf1.idfobjects["Construction".upper()]:
        for fields in items.fieldnames:
            if fields == "key":
                continue
            if fields == "Name":
                continue
            else:
                if items[fields] == "" or items[fields] == "IRTSurface" or items[fields] == "IRTMaterial":
                    continue
                else:
                    items[fields] = items[fields] + f"{replace_string_value}"

    # SAVING THE IDF FILE
    building = str(idf1.idfobjects["Building".upper()][0].Name)
    if base == True:
        idf1.idfobjects["Building".upper()][0].Name = f"BASE{building}{replace_string_value}"
        idf1.saveas(f"{save_folder_path}/BASE{building}{replace_string_value}-converted.idf")
    else:
        idf1.idfobjects["Building".upper()][0].Name = f"{building}{replace_string_value}"
        idf1.saveas(f"{save_folder_path}/{building}{replace_string_value}-converted.idf")
    print(f'{idf1.idfobjects["Building".upper()][0].Name} IDF file is converted and saved...')
    return idf1


def create_combined_idf_archetype(save_folder_path, idflist=list):
    """
    Combines all materials, constructions stored in different idfs and merges them into a single idf file

    Run Order:
    1)Gathers all material and construction objects into lists stored in idfs
    2)Selects the base IDF and deletes all construction and material objects
    3)Creates new materials and constructions from the lists created in 1.
    4)Redefines output variables in a way that BuildME accepts
    5)Saves a new idf file, ready to be used with BuildME

    :param save_folder_path: save folder, where the new merged IDF will be saved
    :param idflist: idf list containing the idf files to be merged
                    this list should only contain an archetype type at a time (e.g. only SFH)
    :return: a new merged archetype idf for BuildME
    """
    cons_list = []
    mat_list = []
    matnomass_list = []
    matwin_list = []
    ffloor_list = []
    cwall_list = []
    matshade_list = []
    matshadecontrol_list = []
    framemat_list = []
    merged_idf = ""

    # Gets required objects in all idf files and puts in a list
    for idf in idflist:
        fhandle = StringIO(idf)
        idf = IDF(fhandle)
        for cons in idf.idfobjects["Construction".upper()]:
            cons_list.append(cons)
        for mat in idf.idfobjects["Material".upper()]:
            mat_list.append(mat)
        for matnomass in idf.idfobjects["Material:NoMass".upper()]:
            matnomass_list.append(matnomass)
        for matwin in idf.idfobjects["WindowMaterial:SimpleGlazingSystem".upper()]:
            matwin_list.append(matwin)
        for matshade in idf.idfobjects["WindowMaterial:Shade".upper()]:
            matshade_list.append(matshade)
        for matshadecontrol in idf.idfobjects["WindowShadingControl".upper()]:
            matshadecontrol_list.append(matshadecontrol)
        for framemat in idf.idfobjects["WindowProperty:FrameAndDivider".upper()]:
            framemat_list.append(framemat)
        for ffloor in idf.idfobjects["Construction:FfactorGroundFloor".upper()]:
            ffloor_list.append(ffloor)
        for cwall in idf.idfobjects["Construction:CfactorUndergroundWall".upper()]:
            cwall_list.append(cwall)
    # Removes duplicate elements

    newobjects = cons_list + matnomass_list + matwin_list + ffloor_list + mat_list + matshade_list + matshadecontrol_list + framemat_list + cwall_list
    reduced_newobjects = reduce(lambda l, x: l.append(x) or l if x not in l else l, newobjects, [])

    # Selects the base IDF, deletes all of the objects and rewrites new objects from other IDFs
    for idf in idflist:
        fhandle = StringIO(idf)
        idf = IDF(fhandle)
        print(f'Now, {idf.idfobjects["Building".upper()][0].Name} values are being written to the merged IDF...')
        if "BASE" in idf.idfobjects["Building".upper()][0].Name:
            idf.removeallidfobjects("CONSTRUCTION")
            idf.removeallidfobjects("MATERIAL")
            idf.removeallidfobjects("MATERIAL:NOMASS")
            idf.removeallidfobjects("WINDOWMATERIAL:SHADE")
            idf.removeallidfobjects("WINDOWSHADINGCONTROL")
            idf.removeallidfobjects("WINDOWPROPERTY:FRAMEANDDIVIDER")
            idf.removeallidfobjects("WINDOWMATERIAL:SIMPLEGLAZINGSYSTEM")
            idf.removeallidfobjects("CONSTRUCTION:FFACTORGROUNDFLOOR")
            idf.removeallidfobjects("CONSTRUCTION:CFACTORUNDERGROUNDWALL")
            # Output Variables needs to be redefined for BuildME
            idf.removeallidfobjects("OUTPUT:VARIABLEDICTIONARY")
            idf.removeallidfobjects("OUTPUT:SURFACES:DRAWING")
            idf.removeallidfobjects("OUTPUT:CONSTRUCTIONS")
            idf.removeallidfobjects("OUTPUT:TABLE:SUMMARYREPORTS")
            idf.removeallidfobjects("OUTPUT:TABLE:MONTHLY")
            idf.removeallidfobjects("OUTPUTCONTROL:TABLE:STYLE")
            idf.removeallidfobjects("OUTPUT:VARIABLE")
            idf.removeallidfobjects("OUTPUT:METER")

            print("Output Variables are being changed...")
            # New output variables are defined by using existing SFH archetype located in data/archetype folder
            SFH_IDF = IDF("..//data//archetype//USA//SFH.idf")
            outputlist = []
            objlist = ["OUTPUT:METER", "OUTPUT:VARIABLEDICTIONARY", "OUTPUT:SURFACES:DRAWING", "OUTPUT:CONSTRUCTIONS",
                       "OUTPUT:TABLE:SUMMARYREPORTS", "OUTPUT:TABLE:MONTHLY", "OUTPUTCONTROL:TABLE:STYLE",
                       "OUTPUT:VARIABLE"]
            for obj in objlist:
                if obj in [x for x in SFH_IDF.idfobjects]:
                    for item in SFH_IDF.idfobjects[f"{obj}"]:
                        outputlist.append(item)
            for newobj in outputlist:
                idf.copyidfobject(newobj)

            # Checks if there are still duplicates
            seen = set()
            dupes = []
            for x in reduced_newobjects:
                if x.Name in seen:
                    dupes.append(x.Name)
                else:
                    seen.add(x.Name)
                    idf.copyidfobject(x)

            # Renames the merged IDF
            building = idf.idfobjects["Building".upper()][0].Name
            building = building.split("-")[0]
            idf.idfobjects["Building".upper()][0].Name = building
            building = building.split("BASE")[1]
            idf.saveas(f"{save_folder_path}/{building}-BuildME.idf")
            merged_idf = idf

    print("The new merged idf file is successfully created")
    return merged_idf


def create_yaml_userreplace(idflist, save_folder_path, Region, Occupation):

    #TODO:NEEDS TO BE UPDATED SIMILAR TO EXCEL ONE
    """
    Creates a yaml file that stores the new values comes from the new idf files.
    The current version now only collects infiltration values.

    :param idflist: list that contains idf files
    :param save_folder_path: Save folder for the yaml file
    :param Region: Region
    :param Occupation: Occupation type
    :return: yaml file with infiltration values
    """

    available_combinations = []
    replace_yaml_list = []
    for idf in idflist:
        fhandle = StringIO(idf)
        idf = IDF(fhandle)
        standard = idf.idfobjects["Building".upper()][0].Name
        standard = standard.split("-")[1]
        standard = standard.split("-RES")[0]
        available_combinations.append(standard)

        # Creating data categories based on the existing structure in replace.xlsx
        keys_tuple = ("Region", "Occupation", "standard", "idfobject", "Name", "objectfield", "Value")

        # Collecting values from several objects
        objlist = ['ZONEINFILTRATION:EFFECTIVELEAKAGEAREA', "ZONEINFILTRATION:DESIGNFLOWRATE",
                   "ZONEINFILTRATION:FLOWCOEFFICIENT"]
        for obj in objlist:
            if obj in [x for x in idf.idfobjects]:
                # Writing all infiltration parameters
                if obj == 'ZONEINFILTRATION:EFFECTIVELEAKAGEAREA':
                    for items in idf.idfobjects["ZoneInfiltration:EffectiveLeakageArea".upper()]:
                        values_tuple = (
                            f"{Region}", f"{Occupation}", f"{standard}", "ZoneInfiltration:EffectiveLeakageArea",
                            f"{items.Name}", "Effective_Air_Leakage_Area", f"{items.Effective_Air_Leakage_Area}")
                        yamldict = dict(zip(keys_tuple, values_tuple))
                        replace_yaml_list.append(yamldict)

                if obj == "ZONEINFILTRATION:DESIGNFLOWRATE":
                    for items in idf.idfobjects["ZoneInfiltration:DesignFlowRate".upper()]:
                        values_tuple = (
                            f"{Region}", f"{Occupation}", f"{standard}", "ZoneInfiltration:DesignFlowRate",
                            f"{items.Name}",
                            "Air_Changes_per_Hour", f"{items.Air_Changes_per_Hour}")
                        yamldict = dict(zip(keys_tuple, values_tuple))
                        replace_yaml_list.append(yamldict)

                if obj == "ZONEINFILTRATION:FLOWCOEFFICIENT":
                    for items in idf.idfobjects["ZoneInfiltration:FlowCoefficient".upper()]:
                        values_tuple = (
                            f"{Region}", f"{Occupation}", f"{standard}", "ZoneInfiltration:FlowCoefficient",
                            f"{items.Name}",
                            "Flow_Coefficient", f"{items.Flow_Coefficient}")
                        yamldict = dict(zip(keys_tuple, values_tuple))
                        replace_yaml_list.append(yamldict)
                        values_tuple = (
                            f"{Region}", f"{Occupation}", f"{standard}", "ZoneInfiltration:FlowCoefficient",
                            f"{items.Name}",
                            "Pressure_Exponent", f"{items.Pressure_Exponent}")
                        yamldict = dict(zip(keys_tuple, values_tuple))
                        replace_yaml_list.append(yamldict)
    # Saving the replace yaml file
    print("Saving the replace yaml file")
    with open(rf'{save_folder_path}\\userreplace.yaml', 'w') as file:
        documents = yaml.dump(replace_yaml_list, file)
    return available_combinations


def create_yaml_usermaterials(merged_idf, save_folder_path, Region):
    """
    Creates two yaml files storing the new materials comes from the new idf files.

    :param merged_idf: Merged IDF file
    :param save_folder_path: Save folder for the yaml file
    :param Region: Region
    :return: a yaml file for nomass materials and another one containing all new materials for odym region assignment
    """

    allmaterials_yaml_list = []
    surrogatematerials_yaml_list = []
    keys_surrogate = (
        "ep_name", "density", "thickness", "conductivity", "resistance (not used in calculations)", "comment",
        "Region")
    keys_allmaterials = ("idf_material", "odym_material", "Region")

    objlist = ['MATERIAL:NOMASS',
               "WINDOWMATERIAL:SHADE", "WINDOWMATERIAL:GAS", "WINDOWMATERIAL:GLAZING",
               "WINDOWMATERIAL:SIMPLEGLAZINGSYSTEM", "MATERIAL:AIRGAP"]
    for obj in objlist:
        if obj in [x for x in merged_idf.idfobjects]:
            if obj == 'MATERIAL:NOMASS':
                for items in merged_idf.idfobjects["Material:NoMass".upper()]:
                    values_surrogate = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                        f"{items.Thermal_Resistance}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

            if obj == "WINDOWMATERIAL:SIMPLEGLAZINGSYSTEM":
                for items in merged_idf.idfobjects["WindowMaterial:SimpleGlazingSystem".upper()]:
                    values_surrogate = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                        f"{items.UFactor}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

            if obj == "WINDOWMATERIAL:GLAZING":
                for items in merged_idf.idfobjects["WindowMaterial:Glazing".upper()]:
                    values_surrogate = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity}",
                        f"{items.Thickness / items.Conductivity}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

            if obj == "WINDOWMATERIAL:GAS":
                for items in merged_idf.idfobjects["WindowMaterial:Gas".upper()]:
                    values_surrogate = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity_Coefficient_A}",
                        f"{items.Thickness / items.Conductivity_Coefficient_A}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

            if obj == "WINDOWMATERIAL:SHADE":
                for items in merged_idf.idfobjects["WindowMaterial:Shade".upper()]:
                    values_surrogate = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity}",
                        f"{items.Thickness / items.Conductivity}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

            if obj == "MATERIAL:AIRGAP":
                for items in merged_idf.idfobjects["Material:AirGap".upper()]:
                    values_surrogate = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                        f"{items.Thermal_Resistance}", "", f"{Region}")
                    values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    surrogate_dict = dict(zip(keys_surrogate, values_surrogate))
                    allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
                    surrogatematerials_yaml_list.append(surrogate_dict)
                    allmaterials_yaml_list.append(allmaterials_dict)

    for items in merged_idf.idfobjects["Material".upper()]:
        values_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
        allmaterials_dict = dict(zip(keys_allmaterials, values_allmaterials))
        allmaterials_yaml_list.append(allmaterials_dict)

    with open(rf'{save_folder_path}\\surrogate_usermaterials.yaml', 'w') as file:
        documents = yaml.dump(surrogatematerials_yaml_list, file)
    with open(rf'{save_folder_path}\\all_usermaterials.yaml', 'w') as file:
        documents = yaml.dump(allmaterials_yaml_list, file)


def colored(r, g, b, text):
    """
    Assigns color to terminal text based on given rgb values
    :param r:
    :param g:
    :param b:
    :param text:
    :return:
    """
    return "\033[38;2;{};{};{}m{} \033[38;2;255;255;255m".format(r, g, b, text)


def update_all_datafile_yaml(idflist, merged_idf, save_folder_path):
    """
    Creates yaml files

    :param idflist: list that contains idf files
    :param merged_idf: Merged IDF file
    :param save_folder_path: save folder for conversion outputs
    :return: summary
    """
    Region = input(colored(0, 255, 255, "Please specify a region name (i.e., USA, UK):"))
    Occupation = input(colored(0, 255, 255, "Please specify the occupancy type (i.e., SFH, MFH, RT):"))
    available_combinations = create_yaml_userreplace(idflist, save_folder_path, Region, Occupation)
    create_yaml_usermaterials(merged_idf, save_folder_path, Region)
    print(" \n")
    print(colored(102, 204, 0,
                  f"SUMMARY:\nThe conversion is completed: The following debugging combinations can be used now:\n{Region}, {Occupation}, {available_combinations}\n\n"))
    print(colored(255, 255, 51,
                  "NEXT STEPS:\n 1. Please update settings.py according to the new idf combinations\n 2. Select a weather file\n 3. Update settings.py by assigning the climate to the region\n 4. Update settings.py by assigning the region to an odym region\n\n"))
    print(colored(255, 0, 0,
                  "CAUTION:\n 1. Please fill missing values in userreplace.yaml, surrogate_usermaterials.yaml and all_usermaterials.yaml files"))


def update_replace_xlsx(idflist, save_folder_path, Region, Occupation):
    """
    Creates an xlsx file that stores the new values comes from the new idf files.
    The current version now only collects infiltration values.

    :param idflist: list that contains idf files
    :param save_folder_path: Save folder for the xlsx file
    :param Region: Region
    :param Occupation: Occupation type
    :return: xlsx file with infiltration values
    """
    replaceme_file_path = "..\\data\\replace.xlsx"
    wb = openpyxl.load_workbook(filename=replaceme_file_path)
    wb.save(f"{save_folder_path}\\replace_updated.xlsx")
    replaceme_file_path = f"{save_folder_path}\\replace_updated.xlsx"
    wb = openpyxl.load_workbook(filename=replaceme_file_path)
    sheet_standard = wb['en-standard']
    wb.create_sheet('cohort')
    sheet_cohort = wb['cohort']
    new_row = ("Region", "Occupation", "cohort", "idfobject",	"Name",	"objectfield",	"Value",	"Comment")
    sheet_cohort.append(new_row)
    available_combinations = []
    std_list = [ "non-standard","standard", "efficient", "ZEB"]
    deletecohort=False


    for idf in idflist:
        fhandle = StringIO(idf)
        idf = IDF(fhandle)
        Standard = idf.idfobjects["Building".upper()][0].Name
        for i in std_list:
            if i in str(Standard):
                deletecohort=True
                standard = Standard.split("-")[1]
                available_combinations.append(standard)

                objlist = ['ZONEINFILTRATION:EFFECTIVELEAKAGEAREA', "ZONEINFILTRATION:DESIGNFLOWRATE",
                           "ZONEINFILTRATION:FLOWCOEFFICIENT","AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK",
                           "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING"]

                for obj in objlist:
                    if obj in [x for x in idf.idfobjects]:
                        # Writing all infiltration parameters
                        if obj == 'ZONEINFILTRATION:EFFECTIVELEAKAGEAREA':
                            for items in idf.idfobjects["ZoneInfiltration:EffectiveLeakageArea".upper()]:
                                new_row = (f"{Region}", f"{Occupation}", f"{standard}", "ZONEINFILTRATION:EFFECTIVELEAKAGEAREA",
                                           f"{items.Name}", "Effective_Air_Leakage_Area", f"{items.Effective_Air_Leakage_Area}")
                                sheet_standard.append(new_row)

                        if obj == "ZONEINFILTRATION:DESIGNFLOWRATE":
                            for items in idf.idfobjects["ZoneInfiltration:DesignFlowRate".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{standard}", "ZONEINFILTRATION:DESIGNFLOWRATE",
                                    f"{items.Name}",
                                    "Air_Changes_per_Hour", f"{items.Air_Changes_per_Hour}")
                                sheet_standard.append(new_row)

                        if obj == "ZONEINFILTRATION:FLOWCOEFFICIENT":
                            for items in idf.idfobjects["ZoneInfiltration:FlowCoefficient".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{standard}", "ZONEINFILTRATION:FLOWCOEFFICIENT",
                                    f"{items.Name}",
                                    "Flow_Coefficient", f"{items.Flow_Coefficient}")
                                sheet_standard.append(new_row)

                        if obj == "AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK":
                            for items in idf.idfobjects["AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{standard}", "AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Coefficient_at_Reference_Conditions", f"{items.Air_Mass_Flow_Coefficient_at_Reference_Conditions}")
                                sheet_standard.append(new_row)

                        if obj == "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING":
                            for items in idf.idfobjects["AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{standard}",
                                    "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Coefficient_When_Opening_is_Closed",
                                    f"{items.Air_Mass_Flow_Coefficient_When_Opening_is_Closed}")
                                sheet_standard.append(new_row)

                            for items in idf.idfobjects["AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{standard}",
                                    "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Exponent_When_Opening_is_Closed",
                                    f"{items.Air_Mass_Flow_Exponent_When_Opening_is_Closed}")
                                sheet_standard.append(new_row)

            else:
                cohort = Standard.split("-")[1]
                available_combinations.append(cohort)

                objlist = [ "LIGHTS","OTHEREQUIPMENT","SHADING:BUILDING:DETAILED",
                           "AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK","AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING"]

                for obj in objlist:
                    if obj in [x for x in idf.idfobjects]:
                        # Writing all infiltration parameters
                        if obj == "LIGHTS":
                            for items in idf.idfobjects["LIGHTS".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}", "LIGHTS",
                                    f"{items.Name}",
                                    "Watts_per_Zone_Floor_Area", f"{items.Watts_per_Zone_Floor_Area}")
                                sheet_cohort.append(new_row)

                        if obj == "OTHEREQUIPMENT":
                            for items in idf.idfobjects["OTHEREQUIPMENT".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}", "OTHEREQUIPMENT",
                                    f"{items.Name}",
                                    "Power_per_Zone_Floor_Area", f"{items.Power_per_Zone_Floor_Area}")
                                sheet_cohort.append(new_row)

                        if obj == "SHADING:BUILDING:DETAILED":
                            for items in idf.idfobjects["SHADING:BUILDING:DETAILED".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}", "SHADING:BUILDING:DETAILED",
                                    f"{items.Name}",
                                    "Transmittance_Schedule_Name", f"{items.Transmittance_Schedule_Name}")
                                sheet_cohort.append(new_row)

                        if obj == "AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK":
                            for items in idf.idfobjects["AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}", "AIRFLOWNETWORK:MULTIZONE:SURFACE:CRACK",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Coefficient_at_Reference_Conditions", f"{items.Air_Mass_Flow_Coefficient_at_Reference_Conditions}")
                                sheet_cohort.append(new_row)

                        if obj == "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING":
                            for items in idf.idfobjects["AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING".upper()]:
                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}",
                                    "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Coefficient_When_Opening_is_Closed",
                                    f"{items.Air_Mass_Flow_Coefficient_When_Opening_is_Closed}")
                                sheet_cohort.append(new_row)

                                new_row = (
                                    f"{Region}", f"{Occupation}", f"{cohort}",
                                    "AIRFLOWNETWORK:MULTIZONE:COMPONENT:DETAILEDOPENING",
                                    f"{items.Name}",
                                    "Air_Mass_Flow_Exponent_When_Opening_is_Closed",
                                    f"{items.Air_Mass_Flow_Exponent_When_Opening_is_Closed}")
                                sheet_cohort.append(new_row)
    if deletecohort==True:
        wb.remove(wb["cohort"])
        available_combinations=reduce(lambda l, x: l.append(x) or l if x not in l else l, available_combinations, [])
        wb.create_sheet('cohort')
        sheet_cohort = wb['cohort']
        new_row = ("Region", "Occupation", "cohort", "idfobject", "Name", "objectfield", "Value", "Comment")
        sheet_cohort.append(new_row)

    wb.save(f"{save_folder_path}\\replace_updated.xlsx")
    return available_combinations


def update_material_xlsx(merged_idf, save_folder_path, Region):
    """
    Creates an xlsx file storing the new materials comes from the new idf files.

    :param merged_idf: Merged IDF file
    :param save_folder_path: Save folder for the xlsx file
    :param Region: Region
    :return: a xlsx file storing nomass materials and all new materials for odym region assignment
    """

    material_file_path = "..\\data\\material_compiled.xlsx"
    wb = openpyxl.load_workbook(filename=material_file_path)
    wb.save(f"{save_folder_path}\\material_updated.xlsx")
    material_file_path = f"{save_folder_path}\\material_updated.xlsx"
    wb = openpyxl.load_workbook(filename=material_file_path)
    sheet_missingmaterials = wb['properties']
    sheet_allmaterials = wb['odym_materials']

    objlist = ['MATERIAL:NOMASS',
               "WINDOWMATERIAL:SHADE", "WINDOWMATERIAL:GAS", "WINDOWMATERIAL:GLAZING",
               "WINDOWMATERIAL:SIMPLEGLAZINGSYSTEM", "MATERIAL:AIRGAP","WINDOWPROPERTY:FRAMEANDDIVIDER"]
    for obj in objlist:
        if obj in [x for x in merged_idf.idfobjects]:
            if obj == 'MATERIAL:NOMASS':
                for items in merged_idf.idfobjects["Material:NoMass".upper()]:
                    new_row_missing = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                       f"{items.Thermal_Resistance}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "WINDOWMATERIAL:SIMPLEGLAZINGSYSTEM":
                for items in merged_idf.idfobjects["WindowMaterial:SimpleGlazingSystem".upper()]:
                    new_row_missing = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                       f"{items.UFactor}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "WINDOWMATERIAL:GLAZING":
                for items in merged_idf.idfobjects["WindowMaterial:Glazing".upper()]:
                    new_row_missing = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity}",
                        f"{items.Thickness / items.Conductivity}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "WINDOWMATERIAL:GAS":
                for items in merged_idf.idfobjects["WindowMaterial:Gas".upper()]:
                    new_row_missing = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity_Coefficient_A}",
                        f"{items.Thickness / items.Conductivity_Coefficient_A}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "WINDOWMATERIAL:SHADE":
                for items in merged_idf.idfobjects["WindowMaterial:Shade".upper()]:
                    new_row_missing = (
                        f"{items.Name}", "SPECIFY DENSITY", f"{items.Thickness}", f"{items.Conductivity}",
                        f"{items.Thickness / items.Conductivity}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "MATERIAL:AIRGAP":
                for items in merged_idf.idfobjects["Material:AirGap".upper()]:
                    new_row_missing = (f"{items.Name}", "SPECIFY DENSITY", "SPECIFY THICKNESS", "SPECIFY CONDUCTIVITY",
                                       f"{items.Thermal_Resistance}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)

            if obj == "WINDOWPROPERTY:FRAMEANDDIVIDER":
                for items in merged_idf.idfobjects["WINDOWPROPERTY:FRAMEANDDIVIDER".upper()]:
                    new_row_missing = (f"{items.Name}", "SPECIFY DENSITY", f"{items.Frame_Width}", f"{items.Frame_Conductance}", f"{items.Frame_Conductance}", f"{Region}")
                    new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
                    sheet_missingmaterials.append(new_row_missing)
                    sheet_allmaterials.append(new_row_allmaterials)



    for items in merged_idf.idfobjects["Material".upper()]:
        new_row_allmaterials = (f"{items.Name}", "SPECIFY ODYM MATERIAL", f"{Region}")
        sheet_allmaterials.append(new_row_allmaterials)
    wb.save(f"{save_folder_path}\\material_updated.xlsx")


def update_all_datafile_xlsx(idflist, merged_idf, save_folder_path):
    """
    Creates xlsx files

    :param idflist: list that contains idf files
    :param merged_idf: Merged IDF file
    :param save_folder_path: save folder for conversion outputs
    :return: summary
    """

    def colored(r, g, b, text):
        return "\033[38;2;{};{};{}m{} \033[38;2;255;255;255m".format(r, g, b, text)

    Region = input(colored(0, 255, 255, "Please specify a region name (i.e., USA, UK):"))
    Occupation = input(colored(0, 255, 255, "Please specify the occupancy type (i.e., SFH, MFH, RT):"))
    available_combinations = update_replace_xlsx(idflist, save_folder_path, Region, Occupation)
    update_material_xlsx(merged_idf, save_folder_path, Region)
    print(" \n")
    print(colored(102, 204, 0,
                  f"SUMMARY:\nThe conversion is completed: The following debugging combinations can be used now:\n{Region}, {Occupation}, {available_combinations}\n\n"))
    print(colored(255, 255, 51,
                  "NEXT STEPS:\n 1. Please update settings.py according to the new idf combinations\n 2. Select a weather file\n 3. Update settings.py by assigning the climate to the region\n 4. Update settings.py by assigning the region to an odym region\n\n"))
    print(colored(255, 0, 0,
                  "CAUTION:\n 1. Please fill missing values in replace_updated.xlsx and material_updated.xlsx files"))


def update_all_datafile_xlsx_gui(idflist, merged_idf, save_folder_path, Region, Occupation):
    """
    Creates xlsx files, same script as update_all_datafile_xlsx but specialized for gui

    :param idflist: list that contains idf files
    :param merged_idf: Merged IDF file
    :param save_folder_path: save folder for conversion outputs
    :param Region: Region name
    :param Occupation: Occupancy type
    :return: summary
    """
    available_combinations = update_replace_xlsx(idflist, save_folder_path, Region, Occupation)
    update_material_xlsx(merged_idf, save_folder_path, Region)
    summarytext = f"The conversion is completed\nThe following debugging combinations can be used now: {available_combinations}"
    nextstepstext = f"1. Please update debugging combinations at settings.py accordingly\n2. Select a weather file for the region\n3. Update settings.py by assigning the climate to the region\n4. Update settings.py by assigning the region to an odym region\n"
    cautiontext = "1. Please fill missing values in replace_updated.xlsx and material_updated.xlsx files"
    return summarytext, nextstepstext, cautiontext, Region, Occupation


def update_all_datafile_yaml_gui(idflist, merged_idf, save_folder_path, Region, Occupation):
    """
    Creates yaml files, same script as update_all_datafile_yaml but specialized for gui

    :param idflist: list that contains idf files
    :param merged_idf: Merged IDF file
    :param save_folder_path: save folder for conversion outputs
    :param Region: Region name
    :param Occupation: Occupancy type
    :return: summary
    """
    available_combinations = create_yaml_userreplace(idflist, save_folder_path, Region, Occupation)
    create_yaml_usermaterials(merged_idf, save_folder_path, Region)
    summarytext = f"The conversion is completed\nThe following debugging combinations can be used now: {available_combinations}"
    nextstepstext = f"1. Please update debugging combinations at settings.py accordingly\n2. Select a weather file for the region\n3. Update settings.py by assigning the climate to the region\n4. Update settings.py by assigning the region to an odym region\n"
    cautiontext = "1. Please fill missing values in replace_updated.xlsx and material_updated.xlsx files"
    return summarytext, nextstepstext, cautiontext, Region, Occupation


if __name__ == "__main__":
    # USER SPECIFIC INPUTS:

    # 2006 IECC idf File Path (standard)
    path1 = "C:\\Users\\sahina\\Downloads\\HOT\\MEX\\sfh-standard-RES0.idf"

    # 2012 IECC idf File Path (efficient)
    path2 = "C:\\Users\\sahina\\Downloads\\HOT\\MEX\\sfh-ZEB-RES0.idf"

    # 2018 IECC File Path (ZEB)
    #path3 = "C:\\Users\\sahina\\Downloads\\HOT\\KUWAIT_VILLA\\Kuwait _Villa_19802010_Reinforced.idf"
    #path4 = "C:\\Users\\sahina\\Downloads\\HOT\\KUWAIT_VILLA\\Kuwait _Villa_Post2010_Reinforced.idf"
    # Main Folder, every generated data will be located
    # Tip: make sure to put the 'r' in front
    folderpath = r"C:\\Users\\sahina\\Downloads\\HOT\\MEX\\New folder"

    convert_idf_to_BuildME(path1, folderpath, replace_string="-en-standard-replaceme-RES-replaceme",
                           replace_string_value="-standard-RES0", base=True)
    convert_idf_to_BuildME(path2, folderpath, replace_string="-en-ZEB-replaceme-RES-replaceme",
                           replace_string_value="-ZEB-RES0", base=False)


    listed = create_combined_idflist(folderpath)
    merged_idf = create_combined_idf_archetype(folderpath, listed)
    update_all_datafile_xlsx(listed, merged_idf, folderpath)
